##Add and remove items are done but still not working properly.
##size varialble in cable encode is acting up.
##Need to grind these two methods. A lot of bug potential.

import configparser as cp
import openpyxl
import datetime
def file_write(sohar,WO,tab,datein='',dateout=None):##Booking set use is DONE.
        '''Adds entry to tab name in store file. Inputs are WO object, datein,
        dateoutand tab name. If datein=='' then use datein from WO object and
        dateout from
        method arguement (Move set). Else if datein!='' then use datein from method argu
        dateout will not be logged (Book set).'''
        if tab=='shistory':
                for k in WO.DH:
                        item=WO.DH[k].copy()
                        item.extend([dateout,WO.base,WO.name])
                        sohar.shistory.append(item)
                for k in WO.cables:
                        item=WO.cables[k].copy()
                        item.extend([dateout,WO.base,WO.name])
                        sohar.shistory.append(item)
                for k in WO.consumables:
                        item=WO.consumables[k].copy()
                        item.extend([None,None,dateout,WO.base,WO.name])
                        sohar.shistory.append(item)

def cable_decode(sn,data):
    
    c1={'0':('Cable, SL-450 (E-Lead) Flat #4/1 AWG, G 5kV','592966','1C code','','','','New'),
        '1':('Cable, SL-450 (E-Lead) Flat #4/1 AWG, G 5kV','592966','1C code','','','','Used'),
        '2':('Cable, SL-450 (E-Lead) Flat #4/1 AWG, SS 5kV', '00870393', '1C590', '', '', '', 'New'),
        '3':('Cable, SL-450 (E-Lead) Flat #4/1 AWG, SS 5kV', '00870393', '1C590', '', '', '', 'Used')}
    c2={'0':('Cable, SL-450 (E-Lead) Flat #2/7 AWG, G 5kV','New','592977'),
        '1':('Cable, SL-450 (E-Lead) Flat #2/7 AWG, G 5kV','Used','592977'),
        '2':('Cable, SL-450 (E-Lead) Flat #2/7 AWG, SS 5kV','New','00870394'),
        '3':('Cable, SL-450 (E-Lead) Flat #2/7 AWG, SS 5kV','Used','00870394')}
    result=[]
    if data[4]=='4':
        for i in range(len(data)):
            if i==4:
                continue
            if data[i]>0:
                a=list(c1[str(i)])
                a[4]=data[i]
                a[3]=sn
                result.append(a)
    elif data[4]=='2':
        for i in range(len(data)):
            if i==4:
                continue
            if data[i]>0:
                a=list(c2[str(i)])
                a[3]=sn
                a[4]=data[0][i]
                result.append(a)
                
    return result    

def cable_encode(data):
    result=[]
    size=''
    x=[]
    c1={('592966','New'):[1,0,0,0],('592966','Used'):[0,1,0,0],
        ('00870393','New'):[0,0,1,0],('00870393','Used'):[0,0,0,1],
        ('592977','New'):[1,0,0,0],('592977','Used'):[0,1,0,0],
        ('00870394','New'):[0,0,1,0],('00870394','Used'):[0,0,0,1]}
    for d in data:
        r=c1[(d[1],d[6])].copy()
        for i in range(len(r)):
            r[i]*=d[4]
        size=d[0][d[0].find('#')+1]
        result.append(r)
    for i in zip(*result):
        x.append(sum(i))
    x.append(size)
    result=x
        
    return result

class Store(object): ## Thinking about adding two more dict's. One for used and one for Booked eqy.
    def __init__(self,name,filename):##Done
        ''' Creates a store object containing 3 dict's: downhole(DH), Cables &
        consumables as attributes.
        Input is Store name (string)Store excel file name. Class reads excel
        file and divides the
        items to 3 dicts. The rules for which item goes to which dictionary are:
        1- If it has no SN then it goes to consumables.
        2- If it has SN, then check first three characters in desciption.
        3- If First 3 c/c's are cab then store in Cables else store in DH.
        4- Fourth dict was added for spare parts in sohar.
        '''
        ##Open Store file and create Store attributes
        
        self.file=filename
        self.name=name

        config=cp.ConfigParser()
        config.read(self.file)

        self.DH=eval(config.get(self.get_name(),"dh"))
        self.cables=eval(config.get(self.get_name(),"cables"))
        self.consumables=eval(config.get(self.get_name(),"consumables"))
        if name=='sohar':
                self.used=eval(config.get(self.get_name(),"used"))
                self.bookings=eval(config.get(self.get_name(),"bookings"))
                self.minStore_map=eval(config.get(self.get_name(),"minstore"))
                self.rogueReels=eval(config.get(self.get_name(),"rogues"))
                self.shistory=eval(config.get(self.get_name(),"shistory"))
                self.rhistory=eval(config.get(self.get_name(),"rhistory"))
                self.chistory=eval(config.get(self.get_name(),"chistory"))
                self.invoices=eval(config.get(self.get_name(),"invoices"))
                self.orders=eval(config.get(self.get_name(),"orders"))
                self.difa=eval(config.get(self.get_name(),"DIFA"))
                self.lekh=eval(config.get(self.get_name(),"LEKH"))
                self.nimr=eval(config.get(self.get_name(),"Nimr"))
                self.jobCount=eval(config.get(self.get_name(),"jobCount"))
                self.difa_cables=eval(config.get(self.get_name(),"difa_cables"))
                self.used_cables=eval(config.get(self.get_name(),"used_cables"))
                self.scrap=eval(config.get(self.get_name(),"scrap"))
        else:
                pass
    def cable_scrap(self,SN,length,arm,date):
        d={'Galv New':'0','Galv Used':'1','SS New':'2','SS Used':'3'}
        if SN in self.cables:
                reel=self.cables[SN]
        elif SN in self.used_cables:
                reel=self.used_cables[SN]
        if reel[int(d[arm])]>=length:
                item=[SN]+reel[:5].copy()+[-length,arm,'','Scrap',date]
                self.scrap[reel[4]+arm[0]][2]+=length
                reel[int(d[arm])]-=length
                self.chistory.append(item)
        else:
                return False
        self.clear_ZQ()
        self.save()
        return True
        
    def clear_difa(self,SN):
        if SN in self.difa:
                self.used[SN]=self.difa[SN].copy()
                del(self.difa[SN])
                self.save()
                return True
        else:
                return False
        

    def update_orders(self,file):
        import openpyxl
        wb=openpyxl.load_workbook(file)
        ws=wb.active
        ords=[]
        for row in ws.rows:
                
                ords.append(list(row))
                
        ords=ords[1:]    
        for row in ords:
                for cell in row:
                        row[row.index(cell)]=cell.value
                if row[3]!=None:
                        row[3]=row[3].strftime('%d-%b-%Y')

                self.orders[str(row[0])]=row[1:]
        print(self.orders)
                
        
        self.save()
        return
        
                    
    def create_wo(self,name,base,sn_list,pn_list):
        name=name.upper()
        pump400={'02025831': ['Coupling, ESP 17mm INV 15T x 17mm INV 15T LG:73.0mm Monel', '02025831', '1C487', None, 55], '02029183': ['Coupling, ESP 17mm INV 15T x 22mm INV 20T LG:73.0mm Monel', '02029183', '1C488', None, 47]}
        pump500={'2031612': ['Coupling, ESP 17mm INV 15T x 25mm INV 24T LG: 73mm WN-311', '2031612', '1C489', None, 39], '02024923': ['Coupling, ESP 20mm INV 18T x 20mm INV 18T LG: 75mm CS', '02024923', '1C490', None, 4]}
        pump300={'a':16,'b':17}
        motor400={'c':22,'d':23}
        motor500={'2000304#': ['O-Ring, 072-078-36-2-(AF-100) GOST 9833/GOST 18829', '20003047', '1C546', None, 461]}
        motor300={'c':26,'d':27}
        seal300={'e':32,'f':33}
        seal400={'e':34,'f':35}
        seal500={'20003047': ['O-Ring, 072-078-36-2-(AF-100) GOST 9833/GOST 18829', '20003047', '1C546', None, 10]}
        splice={'719328': ['Tape, HI-TEMP sintered and extruded w/liner 1in x 18yds', '719328', '1C579', None, 2]}
        pumpused={}
        motorused={}
        sealused={}
        pcount=0
        mcount=0
        scount=0
        dh={}
        cables={}
        cons={}
        internal={}
        rogue=[]
        status='Unbooked'
        sent=''
        
        if len(sn_list)==0 and len(pn_list)==0:
                config=cp.ConfigParser()
                config.read(self.file)
                config.add_section(str(name))
                config.set(str(name),"dh",str(dh))
                config.set(str(name),"cables",str(cables))
                config.set(str(name),"consumables",str(cons))
                config.set(str(name),"rogue",str(rogue))
                config.set(str(name),"base",str(base))
                config.set(str(name),"status",str(status))
                config.set(str(name),"sent",str(sent))
                config.set(str(name),"internal",str(internal))
                config.set(str(name),"name",str(name.upper()))
                f=open(self.file,'w')
                config.write(f)
                f.close()
                
        for s in sn_list:
                if s[0] in self.DH and s[1]<=self.DH[s[0]][4]:
                        dh[s[0]]=self.DH[s[0]].copy()
                        dh[s[0]][4]=s[1]
                elif s[0] in self.cables and s[1]<=sum(self.cables[s[0]][:4]):
                        a=cable_decode(s[0],self.cables[s[0]])
##                        cables[s[0]]=self.cables[s[0]].copy()
##                        cables[s[0]][4]=s[1]
                        for i in range(len(a)):
                                if i>0:
                                        cables[s[0]+'/'+str(i+1)]=a[i]
                                else:
                                        cables[s[0]]=a[i]
        
        for p in pn_list:
                if len(p)>0:
                        if p[0] in self.consumables and p[1]<=self.consumables[p[0]][4]:
                                cons[p[0]]=self.consumables[p[0]].copy()
                                cons[p[0]][4]=p[1]
                else:
                        pass
        
##        for c in cables:
##                if c in self.rogueReels:
##                        rogue=self.rogueReels[c].copy()
##                else:
##                        rogue=[]
        for k in dh:
                if 'Pump,' in dh[k][0]:
                        pcount+=1
                        if 'B 538' in dh[k][0]:
                                pumpused=pump500
                        elif 'B 400' in dh[k][0]:
                                pumpused=pump400
                        elif '338' in dh[k][0]:
                                pumpused=pump300
                        else:
                                pumpused={}
                        
                elif 'Motor,' in dh[k][0]:
                        mcount+=1
                        if 'B 562' in dh[k][0]:
                                motorused=motor500
                        elif 'B 456' in dh[k][0]:
                                motorused=motor400
                        elif '375' in dh[k][0]:
                                motorused=motor300
                        else:
                                motorused={}
                elif 'Motor Seal,' in dh[k][0]:
                        scount+=1
                        if 'B 538' in dh[k][0]:
                                sealused=seal500
                        elif 'B 400' in dh[k][0]:
                                sealused=seal400
                        elif '338' in dh[k][0]:
                                sealused=seal300
                        else:
                                sealused={}                       
        for k in pumpused:
                pumpused[k][4]*=pcount
        for k in motorused:
                motorused[k][4]*=mcount
        for k in sealused:
                sealused[k][4]*=scount
        internal.update(pumpused)
        internal.update(motorused)
        internal.update(sealused)
        if len(cables)>0:
                internal.update(splice)
                 
        
        config=cp.ConfigParser()
        config.read(self.file)
        config.add_section(str(name))
        config.set(str(name),"dh",str(dh))
        config.set(str(name),"cables",str(cables))
        config.set(str(name),"consumables",str(cons))
        config.set(str(name),"rogue",str(rogue))
        config.set(str(name),"base",str(base))
        config.set(str(name),"status",str(status))
        config.set(str(name),"sent",str(sent))
        config.set(str(name),"internal",str(internal))
        config.set(str(name),"name",str(name.upper()))
        f=open(self.file,'w')
        config.write(f)
        f.close()
        return [pumpused,motorused,sealused,splice]
        
    def get_wos(self):
        config=cp.ConfigParser()
        config.read(self.file)
        w=[]
        for section in config.sections():
                if section!='sohar':
                        w.append(section)
        return w
                
    def get_name(self):##Done
        return self.name

    def find_item(self,ID,IDtype='SN'):##Done
        ''' Searches for item in excel file. Input is ID (SN or PN). IDtype
        is used to specify the ID is SN or PN. Default is set to SN. Returns
        a Bool, the sheet name and quantity if Bool is True'''
        if IDtype=='SN':##Check ID type
                if ID in self.DH:                                               ##If SN is used check if it is in DH dict
                        return True, self.DH[ID][4]                        ##Return True and QTY is Yes
                elif ID in self.cables:                                         ##If not in DH check if it is in cable dict
                        return True, sum(self.cables[ID][:4])                   ##Return True and QTY is Yes
                
                else:
                        return False,''                                         ##If not found return false and empty qty
        else:
                if ID in self.consumables:                                      ##If ID type is not SN check if PN is in consumables and sp dict
                        return True, self.consumables[ID][4]             ##Return true and QTY if yes
                else:
                        
                        return False, ''                         ##Return False and QTY as '' if NO

    def sub_WO(self,WO): ##DONE
        '''Subtracts the WO from copy of master store(Sohar). Sohar & WO dict's should be copied and the
        above operations to be carried out on copies.'''
        ##check WO availability
        s,a=WO.is_avail(self)
        if s:
        ## Yes: Loop over each item in WO dicts and subtract it from sohar dicts
                for k in WO.DH:
                        self.DH[k][4]-=WO.DH[k][4]
                for k in WO.cables:
                        if '/' in k:
                                continue
                        del(self.cables[k])
                for k in WO.consumables:
                        self.consumables[k][4]-=WO.consumables[k][4]
                return True
        else:
##                print('Not available', a)
                return False

    def add_WO(self,WO):#DONE
            
        '''Adds WO back to Sohar. takes WO object and condition '''
        ##For each dict in WO; check to see if each item is in sohar store dicts
        c=[]
        for k in WO.DH:
                if True not in self.find_item(k,'SN'):##check to see if each item is not in sohar DH dict
                        self.DH[k]=WO.DH[k].copy() ##No:Add item to sohar.DH
                else:
                        pass
##                        print('Item already in store', k)
        for k in WO.cables:
                sn=WO.cables[k][3]
                c.append(WO.cables[k])
        if len(WO.cables)>0:
                self.cables[sn]=cable_encode(c)
##                if '/' in k:
##                        continue
##                if True not in self.find_item(k,'SN'):##check to see if each item is not in sohar cables dict
##                        self.cables[k]=WO.cables[k].copy()##No:Add item to sohar.cables
##                                    
##                else:
##                        pass
##                        print('Item already in store', k) ##Yes: Return False
        for k in WO.consumables:
                if True not in self.find_item(k,'PN'): ##check to see if each item is not in sohar cons dict
                        self.consumables[k]=WO.consumables[k].copy()##No: add new entry for this item
                                    
                else:
                        self.consumables[k][4]+=WO.consumables[k][4] ## Yes: increment quantity only
       

    def clear_ZQ(self): ## Done
        '''checks the store copy for any items with zero quantities and removes them
        from dictionary. Then writes to sohar file?'''
        D=[self.DH]
        zeros=[]
        for dicts in D:                                                         ##Loop over each dict in D.
                for k in dicts:                                                 ## Loop over keys in D:
                        if dicts[k][4]==0:                                      ##check if qty of item k is zero
                                zeros.append(k)                                 ##Append key in list zeros
                for z in zeros:                                                 ##After loop ends use zeros to delete recorded keys from dictionary
                        del(dicts[z])
                zeros=[]
        for k in self.consumables:                                              ## Loop over keys in consumables:
                if self.consumables[k][4]==0:                                   ##check if qty of item k is zero
                        zeros.append(k)                                         ##Append key in list zeros
        for z in zeros:
                del(self.consumables[z])                                        ##After loop ends use zeros to delete recorded keys from consumables
        
        
    def return_cable(self,cables,enc,date,Base,well):##marking
            ##cable is a dict of decoded cables.
        
        for c in cables:
                if c[6]=='Used':
                        for k in enc:
                                self.difa_cables[k]=enc[k].copy()
                                self.difa_cables[k].extend([Base,well])
                elif c[6]=='New':
                        self.cables.update(enc)

                
                a=c.copy()
                a.extend(['',Base,well])
                a[5]=date
                print(a)
                self.shistory.append(a)
        if Base=='Lekhwair':
                del(self.lekh[c[3]])
        elif Base=='Nimr':
                del(self.nimr[c[3]])
        return            
    def return_item(self,data,date,Base):
        ##Data is a list of SN's.
        for entry in data:
                if Base=='Lekhwair':

                        if entry[0] in self.lekh:
                                item=self.lekh[entry[0]].copy()
                                item[4]=entry[1]
                                item[6]=entry[2]
                                if item[6]=='Used':
                                        self.difa[entry[0]]=item.copy()
                                        self.difa[entry[0]].extend([Base,entry[3]])
                                elif item[6]=='New':
                                        if item[0][:3]=='Cab':
                                                self.cables[entry[0]]=item.copy()
                                        elif item[0][:3]!='Cab':
                                                self.DH[entry[0]]=item.copy()
                                                
                                item[5]=date
                                item.extend(['',Base,entry[3]])
                                self.shistory.append(item)
                                del(self.lekh[entry[0]])
                        else:

                                return False
                
                elif Base=='Nimr':

                        if entry[0] in self.nimr:
                                item=self.nimr[entry[0]].copy()
                                item[4]=entry[1]
                                item[6]=entry[2]
                                if item[6]=='Used':
                                        self.difa[entry[0]]=item.copy()
                                        self.difa[entry[0]].extend([Base,entry[3]])
                                elif item[6]=='New':
                                        if item[0][:3]=='Cab':
                                                self.cables[entry[0]]=item.copy()
                                        elif item[0][:3]!='Cab':
                                                self.DH[entry[0]]=item.copy()
                                item[5]=date
                                item.extend(['',Base,entry[3]])
                                self.shistory.append(item)
                                del(self.nimr[entry[0]])
                        else:

                                return False
        
    def add2base(self,WO):
        c=[]
        if WO.get_base()=='Lekhwair':
                for k in WO.DH:
                        self.lekh[k]=WO.DH[k].copy()
                for k in WO.cables:
                        c.append(WO.cables[k])
                        sn=WO.cables[k][3]
                if len(WO.cables)>0:
                        self.lekh[sn]=cable_encode(c)
        elif WO.get_base()=='Nimr':
                
                for k in WO.DH:
                        self.nimr[k]=WO.DH[k].copy()
                for k in WO.cables:
                        c.append(WO.cables[k])
                        sn=WO.cables[k][3]
                if len(WO.cables)>0: 
                        self.nimr[sn]=cable_encode(c)
        
    def Move_set(self,WO,dateout=''): ##DONE
        '''If WO.bookng is false, Subtracts WO from sohar copy.Copy WO to sohar
        history. Retrieve
        the date in from sohar store to be added in history. Copy dateout
        to sohar history.
        If WO.booking is True then move set from Booked tab to sohar history.'''
        if WO.get_status()=='Booked': ## If set is booked
                WO.unbook_set(self) ##Add set back to sohar and delete it from Booked tab
                self.sub_WO(WO) ## Subtract it again from sohar redundant but necessary
                self.clear_ZQ() ## Clear any zeroes in sohar store.
                self.add2base(WO)
                file_write(self,WO,'shistory','',dateout)
                self.save()    ##Save the dict to file
                WO.set_status('Sent',dateout) ##CHange WO status to sent
                WO.sent=dateout
        elif WO.get_status()=='Unbooked': ## If not booked
                self.sub_WO(WO)  ## Subtract WO from sohar store
                self.clear_ZQ() ## Clear any zeroes in sohar store.
                self.add2base(WO)
                file_write(self,WO,'shistory','',dateout)
                self.save()    ##Save the dict to file
                WO.set_status('Sent',dateout)  ##CHange WO status to sent
                WO.sent=dateout
        else:
                return 'Already sent'
        


    def save(self): ##Done
        '''Saves all changes by saving copies to original store
        Writes sohar original dict to excel file. '''
        ## Update sohar original with copy of store.
        ##opens excel
        config=cp.ConfigParser()
        config.read('config.ini')
        config.set('sohar','dh',str(self.DH))
        config.set('sohar','cables',str(self.cables))
        config.set('sohar','consumables',str(self.consumables))
        config.set('sohar','used',str(self.used))
        config.set('sohar','bookings',str(self.bookings))
        config.set('sohar','shistory',str(self.shistory))
        config.set('sohar','rhistory',str(self.rhistory))
        config.set('sohar','chistory',str(self.chistory))
        config.set('sohar','minstore',str(self.minStore_map))
        config.set('sohar','rogues',str(self.rogueReels))
        config.set('sohar','invoices',str(self.invoices))
        config.set('sohar','orders',str(self.orders))
        config.set('sohar','DIFA',str(self.difa))
        config.set('sohar','LEKH',str(self.lekh))
        config.set('sohar','Nimr',str(self.nimr))
        config.set('sohar','jobCount',str(self.jobCount))
        config.set('sohar','difa_cables',str(self.difa_cables))
        config.set('sohar','used_cables',str(self.used_cables))
        config.set('sohar','scrap',str(self.scrap))
        f=open(self.file,'w')
        config.write(f)
        f.close()
        
        return
       
    def update_reel(self,SN1,SN2,length,armor,well,date): ## Done, but need to think about how to represent adding cables with different armors (Rogue reels)
        '''method to change cable reels to fit WO creation. SN1 is the reel to
        receive extra cable. SN2 is the reel to deduct cable from. length is
        the length of cable to be added and removed. Method should find SN2 in
        sohar cabel dict, and deduct the length to the qty, if it checks out
        it should find SN1 and add the length to it. The aim to make the cable
        available in sohar store when WO is requested'''
        arm=['Galv New','Galv Used','SS New','SS Used']
        index=arm.index(armor)
        if self.cables[SN2][index]>=length:
                x1=self.cables[SN1].copy()
                x2=self.cables[SN2].copy()
                self.cables[SN1][index]+=length
                self.cables[SN2][index]-=length
                a=[SN1]+x1+[length,armor,well,'Addition',date]
                b=[SN2]+x2+[-length,armor,well,'Removal',date]
                self.chistory.append(a)
                self.chistory.append(b)
        else:
                return False
##        metal_ch1='SS' in self.cables[SN1][0]
##        metal_ch2='SS' in self.cables[SN2][0]
##        if metal_ch1!=metal_ch2:
##                self.rogueReels[SN1]=[length,SN2,SN1]
##                
##        if self.cables[SN2][4]>=length:                         ##Check if length<=qty in reel SN2
##                
##                a=self.cables[SN1][0:5].copy()
##                a.extend([self.cables[SN1][4]+length,length,SN2,None])
##                b=self.cables[SN2][0:5].copy()
##                b.extend([self.cables[SN2][4]-length,-length,None,SN1])
##                
##                self.chistory.append(a)
##                self.chistory.append(b)          
##                self.cables[SN1][4]+=length                     ##Yes: increment qty of SN1 by length
##                self.cables[SN2][4]-=length                     ## decrement qty of SN2 by length
##                self.save()
##                return True
##        else:
##                return False
        
    def get_booked(self): ##DONE
        ''''Returns list with name of booked WO's.'''
        pass

    def repair(self,SN,SP,date):## DONE
        ''' Moves a DH or cable from used tab to sohar tab and store.
        takes SN of eqy to be repaired and list of tuples for spare part PN's & QTY as input [(PN1,QTY1),(PN2,QTY2)].
        Searches for SN in used tab. Returns list
        [desc,pn,sn,qty,datein,condition='used']. It then adds this item
        to sohar's appropriarte dict and then saves sohar. It also deducts spare parts used from sohar.sp'''
        
        r=0
        item=[]
        
        if SN in self.used: ##Check if item is in used store
                item.append(self.used[SN].copy())
                item[0]=item[0][:7]
                item[0].append(date)
                for p in SP:
                        a=self.consumables[p[0]].copy()
                        a[4]=p[1]
                        item.append(a)

                for p in SP: ## Check if there is enough spare parts
                        if str(p[0]) not in self.consumables or self.consumables[p[0]][4]<p[1]:
##                                print('Insufficient qty from spare part PN: '+str(p[0]))
                                return 'Insufficient',str(p[0])
                if self.used[SN][0][:3]=='Cab': ##check if item is cable
                        self.cables[SN]=self.used[SN].copy() ##Add item to cable
                        self.cables[SN][6]='Repaired'
                        self.cables[SN]=self.cables[SN][:7]
                        del(self.used[SN]) ##delete item from used dict
                else:
                        self.DH[SN]=self.used[SN].copy() ## add item to DH
                        self.DH[SN][6]='Repaired'
                        self.DH[SN]=self.DH[SN][:7]
                        del(self.used[SN]) ##delete item from used dict
                for p in SP: ##deduct spares to be used
                        self.consumables[p[0]][4]-=int(p[1])
        elif SN in self.used_cables:
                item=self.used_cables[SN].copy()
                item=item[:-2]
                for p in SP: ## Check if there is enough spare parts
                        if str(p[0]) not in self.consumables or self.consumables[p[0]][4]<p[1]:
##                                print('Insufficient qty from spare part PN: '+str(p[0]))
                                return 'Insufficient',str(p[0])
                        
                for p in SP: ##deduct spares to be used
                        self.consumables[p[0]][4]-=int(p[1])
                self.cables[SN]=item.copy()
                del(self.used_cables[SN])
                a=[SN]+item.copy()+[0,'','','Repaired',date]
                self.chistory.append(a)
                        
                                
                
        else:                              
##                print('Item not in used store.')
                return 'Itemnot','na'

        
        self.rhistory[(SN,date)]=item
        self.clear_ZQ()
        self.save()
        return 'OK',''
     
    def new_batch(self,filename,datein): ##DONE
        ''' Method for adding new batch of incoming shipment to sohar store.
        The input is excel file name and the incoming date of the batch.'''

        
        inv=[]
        ##Open batch file & read rows.
        import openpyxl
        wb=openpyxl.load_workbook(filename)
        ws=wb.active
        batch=[]
        for row in ws.rows:
                batch.append(list(row))
        for row in batch:
                for c in row:
                        row[row.index(c)]=c.value
        batch=batch[1:]
        
        
        ##Check if all rows are not in sohar store. Save rows found in sohar
        check=[]
        for row in batch:
                if row[3]!=None:
                        s,q=self.find_item(str(row[3]),'SN')
                        if s:
##                                print('Item already in Sohar. Please check SN.')
                                check.append(str(row[3]))

        for row in batch: ##Loop to turn non None SN's & PNs to strings instead of numbers
                row[1]=str(row[1])
                row[2]=str(row[2])
                if row[3]!= None:
                        row[3]=str(row[3])                        
        
        ##Add items to sohar dicts:        
        for row in batch:
                if row[3]!=None and row[0][:3]!='Cab' and row[3] not in check:
                        del(row[-1])
                        row.extend([datein,'New'])
                        self.DH[row[3]]=row
                elif row[3]!=None and row[0][:3]=='Cab' and row[3] not in check:
                        del(row[-1])
                        row.extend([datein,'New'])
                        self.cables[row[3]]=cable_encode([row])
                elif row[3]==None and row[-1] not in self.invoices :
                        if row[1] in self.consumables:
##                                print(self.consumables[row[1]])
                                self.consumables[row[1]][4]+=row[4]
                        else:
                                
                                self.consumables[row[1]]=row[:-1]
                
                elif row[3]==None and row[-1] in self.invoices:
                        inv.append(row[-1])
                        
        
        for i in ws['F']:
                self.invoices.add(i.value)
        
        
        ##Add batch to incoming sheet. Can be used to validate inventory in future.

        config=cp.ConfigParser()
        config.read(self.file)
        config.set('sohar','incoming',str(batch))
        f=open(self.file,'w')
        config.write(f)
        f.close()
        self.save()
        return check,inv

    def search_history(self,SN):##DONE
        '''Searches the sohar history tab for a specific SN. Returns a list of all rows containing this item
        Input is SN string.'''

        seares=[]
        for row in self.shistory:
                if row[3]==SN:
                        seares.append(row)
        
        return seares

    def av_pn_map(self): ##DONE
        '''Creates dictionary for all items in sohar store. The key is PN and value is qty.
        It is an accumulation of all qty's for each PN in the entire store.'''
        pnMap={}
        for k in self.DH:
                if self.DH[k][1] in pnMap:
                        pnMap[self.DH[k][1]]+=self.DH[k][4]
                else:
                        pnMap[self.DH[k][1]]=self.DH[k][4]
        for k in self.cables:
                if self.cables[k][1] in pnMap:
                        pnMap[self.cables[k][1]]+=self.cables[k][4]
                else:
                        pnMap[self.cables[k][1]]=self.cables[k][4]
        for k in self.consumables:
                if k in pnMap:
                        pnMap[k]+=self.consumables[k][4]
                else:
                        pnMap[k]=self.consumables[k][4]
        return pnMap
        
    def used_pn_map(self): ## DONE
        '''Creates dictionary for all items in sohar used store. The key is PN and value is qty.
        It is an accumulation of all qty's for each PN in the entire store.'''
        pnMap={}
        for k in self.used:
                if self.used[k][1] in pnMap:
                        pnMap[self.used[k][1]]+=self.used[k][4]
                else:
                        pnMap[self.used[k][1]]=self.used[k][4]
        return pnMap
    def booked_pn_map(self):##DONE
        '''Creates dictionary for all items in sohar booked store. The key is PN and value is qty.
        It is an accumulation of all qty's for each PN in the entire store.'''
        pnMap={}
        for row in self.bookings:
                if row[1] in pnMap:
                        pnMap[row[1]]+=row[4]
                else:
                        pnMap[row[1]]=row[4]
        return pnMap
    def inc_pn_map(self): #DONE
        '''Creates dictionary for all items in sohar incoming sheet. The key is PN and value is qty.
        It is an accumulation of all qty's for each PN that ever entered sohar.'''
        pnMap={}
        pn=[]
        qty=[]
        config=cp.ConfigParser()
        config.read(self.file)
        incoming=eval(config.get("sohar",'incoming'))

        for row in incoming:
                pn.append(row[1])
                qty.append(row[4])
                
        for p,q in zip(pn,qty):
                if p in pnMap:
                        pnMap[p]+=q
                else:
                        pnMap[p]=q
        return pnMap
    def search_pn(self,pn):##DONE, proud of this one ;)
        '''Searches sohar history tab for a specific PN to determine if the item was returned or still out of sohar.'''

        sn=[]
        stateList=[]
        state=None
        for row in self.shistory: ##loop over shistory and accumulate all SN's for the input PN that can be found in history
                if row[1]==str(pn): ## check if pn we are searching for is in the given list
                        sn.append(row[3]) ## if yes append the SN to list
        sn=list(set(sn)) ##eliminate duplicates from list
                        
        for s in sn: ##For each SN loop over the file and check what is the final state of this item
                for row in self.shistory:
                        if row[3]==str(s): ## check is SN we are searching for is in this cell.
                                if row[7]=='' or row[7]==None: ##if yes, check to see if for this entry date out has a date or not.
                                        state=0 ##If it does not have a date that means for this entry the the item was returned to sohar
                                else:
                                        state=1 ##if it does have a date that means that the item was sent out of sohar
                stateList.append(state) ##The final state will be appended to the state list
        return sum(stateList) ## The sum of ones represents the number of items for this PN that where sent and not returned to sohar.        

     
            
    def min_store(self,pn,qty):
        '''Create sohar.minstore attribute dict. It contains a map of PN: min qty required. Input is
        pn(string) and qty (int), Returns dictionary mapping PN to min qty for store. '''
        
        self.minStore_map[pn]=qty            

    def generate_report(self,file):
        import openpyxl
        wb=openpyxl.Workbook()
##        wb2=openpyxl.load_workbook(self.file)
        wb.create_sheet('Downhole Equipment',0)
        wb.create_sheet('Cables',1)
        wb.create_sheet('Consumables',2)
        wb.create_sheet('Used Equipment',3)
        wb.create_sheet('Used cables',4)
        wb.create_sheet('Booked Sets',5)
        wb.create_sheet('Sohar History',6)
        wb.create_sheet('Repair History',7)
        wb.create_sheet('Cable History',8)
        wb.create_sheet('DIFA equipment',9)
        wb.create_sheet('DIFA cables',10)


        ws=wb['Downhole Equipment']
        cols1=['B','C','D','E','F','G','H']
        
        header=['Type','Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition']
        ws.append(header)
        for k in self.DH: ##  Needto add all types of downhole equipment
                if self.DH[k][0][:3]=='Pum':
                        ws['A'+str(ws.max_row+1)]='Pump'
                        for i in cols1:
                                ws[i+str(ws.max_row)]=self.DH[k][cols1.index(i)]
                elif self.DH[k][0][:7]=='Motor, ':
                        ws['A'+str(ws.max_row+1)]='Motor'
                        for i in cols1:
                                ws[i+str(ws.max_row)]=self.DH[k][cols1.index(i)]
                elif self.DH[k][0][:7]=='Motor S':
                        ws['A'+str(ws.max_row+1)]='Seal'
                        for i in cols1:
                                ws[i+str(ws.max_row)]=self.DH[k][cols1.index(i)]
                elif self.DH[k][0][:3]=='Gas':
                        ws['A'+str(ws.max_row+1)]='Gas Sep'
                        for i in cols1:
                                ws[i+str(ws.max_row)]=self.DH[k][cols1.index(i)]
                elif self.DH[k][0][:3]=='Sen':
                        ws['A'+str(ws.max_row+1)]='Sensor'
                        for i in cols1:
                                ws[i+str(ws.max_row)]=self.DH[k][cols1.index(i)]
                elif self.DH[k][0][:3]=='MLE':
                        ws['A'+str(ws.max_row+1)]='MLE'
                        for i in cols1:
                                ws[i+str(ws.max_row)]=self.DH[k][cols1.index(i)]
        ws=wb['Cables']
        ws.append(['Reel no','Galv New','Galv Used','SS New','SS Used','Size'])
        for k in self.cables: ##  Needto add all types of downhole equipment
                ws['A'+str(ws.max_row+1)]=k
                for i in cols1[:5]:
                        ws[i+str(ws.max_row)]=self.cables[k][cols1.index(i)]
        ws=wb['Consumables']
        ws.append(['Type','Description','Part Number','1C Code','Serial Number','Quantity'])
        for k in self.consumables:
                if self.consumables[k][0][:4]=='O-ri' or self.consumables[k][0][:4]=='O-Ri':
                        ws['A'+str(ws.max_row+1)]='O-ring'
                        for i in cols1[:5]:
                                ws[i+str(ws.max_row)]=self.consumables[k][cols1[:5].index(i)]
                        
                elif self.consumables[k][0][:4]=='Coup':
                        ws['A'+str(ws.max_row+1)]='Coupling'
                        for i in cols1[:5]:
                                ws[i+str(ws.max_row)]=self.consumables[k][cols1[:5].index(i)]
                        
                elif self.consumables[k][0][:4]=='Slee':
                        ws['A'+str(ws.max_row+1)]='Sleeve'
                        for i in cols1[:5]:
                                ws[i+str(ws.max_row)]=self.consumables[k][cols1[:5].index(i)]
                       
                elif self.consumables[k][0][:4]=='Tape':
                        ws['A'+str(ws.max_row+1)]='Tape'
                        for i in cols1[:5]:
                                ws[i+str(ws.max_row)]=self.consumables[k][cols1[:5].index(i)]
                        
                elif self.consumables[k][0][:4]=='Valv':
                        ws['A'+str(ws.max_row+1)]='Valve'
                        for i in cols1[:5]:
                                ws[i+str(ws.max_row)]=self.consumables[k][cols1[:5].index(i)]
                        
                elif self.consumables[k][0][:4]=='Gask':
                        ws['A'+str(ws.max_row+1)]='Gasket'
                        for i in cols1[:5]:
                                ws[i+str(ws.max_row)]=self.consumables[k][cols1[:5].index(i)]
                        
                else:
                        ws['A'+str(ws.max_row+1)]='Consumables'
                        for i in cols1[:5]:
                                ws[i+str(ws.max_row)]=self.consumables[k][cols1[:5].index(i)]
##                        ws['D'+str(ws.max_row)]=self.consumables[k][-1]
       

        ws=wb['Used Equipment']
        ws.append(['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition','Base','Well'])
        for k in self.used:
                ws.append(self.used[k])
        ws=wb['Used cables']
        ws.append(['Reel no','Galv New','Galv Used','SS New','SS Used','Size'])
        for k in self.used_cables: ##  Needto add all types of downhole equipment
                ws['A'+str(ws.max_row+1)]=k
                for i in cols1[:5]:
                        ws[i+str(ws.max_row)]=self.cables[k][cols1.index(i)]

        ws=wb['DIFA equipment']
        ws.append(['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition','Base','Well'])
        for k in self.difa:
                ws.append(self.difa[k])

        ws=wb['DIFA cables']
        ws.append(['Reel no','Galv New','Galv Used','SS New','SS Used','Size','Base','Well'])
        for k in self.difa_cables: ##  Needto add all types of downhole equipment
                ws['A'+str(ws.max_row+1)]=k
                for i in cols1:
                        ws[i+str(ws.max_row)]=self.difa_cables[k][cols1.index(i)]
        

        ws=wb['Booked Sets']
        ws.append(['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition','Base','Well'])
        for row in self.bookings:
                ws.append(row)

        ws=wb['Sohar History']
        ws.append(['Description','Part Number','1C Code','Serial number','Quantity','Date In','Condition','Date out','Base','Well'])
        for row in self.shistory:
                ws.append(row)

        ws=wb['Repair History']
        ws.append(['Description','Part Number','1C Code','Serial number','Quantity','Date In','Condition','Repair Date'])
        for k in self.rhistory:
                for L in self.rhistory[k]:
                        ws.append(L)

        ws=wb['Cable History']
        ws.append(['Reel no','Galv New','Galv Used','SS New','SS Used','Size','Update','Arm/Cond','Well','Operation','Date'])
        for row in self.chistory:
                ws.append(row)
        
                
                
        for s in ['Downhole Equipment', 'Cables','Consumables', 'Used Equipment','Booked Sets','Cable History','Sohar History','Repair History','DIFA equipment',
                  'DIFA cables','Used cables']:
                ws=wb[s]
                bd = openpyxl.styles.Side(style='thick', color="000000")
                bd2 = openpyxl.styles.Side(style='thin', color="000000")
                ##Header formatting
                for c in ws['1']:
                        c.font=openpyxl.styles.Font(size=12,bold=True)
                        c.fill=openpyxl.styles.PatternFill(fill_type='solid',start_color='FF0000',end_color='FF0000')
                        c.border=openpyxl.styles.Border(left=bd,right=bd,top=bd,bottom=bd)
                        c.alignment=openpyxl.styles.Alignment(horizontal='center',vertical='center')
                ##Body formatting
                for i in range(2,ws.max_row+1):
                        for c in ws[str(i)]:
                                c.fill=openpyxl.styles.PatternFill(fill_type='solid',start_color='00FFFF',end_color='FF0000')
                                c.border=openpyxl.styles.Border(left=bd2,right=bd2,top=bd2,bottom=bd2)
                                c.alignment=openpyxl.styles.Alignment(horizontal='center',vertical='center')
                ## cell dimensions
                if ws.title in['Cables','Used cables','DIFA cables','Cable History']:
                        ws.row_dimensions[1].height=30
                        ws.column_dimensions['A'].width=13
                        ws.column_dimensions['B'].width=13
                        ws.column_dimensions['C'].width=13
                        ws.column_dimensions['D'].width=13
                        ws.column_dimensions['E'].width=13
                        ws.column_dimensions['F'].width=13
                        ws.column_dimensions['G'].width=13
                        ws.column_dimensions['H'].width=13
                        ws.column_dimensions['I'].width=13
                        ws.column_dimensions['J'].width=13
                        ws.column_dimensions['K'].width=13
                        ws.freeze_panes = 'A2'
                        
                elif ws.title in {'Used Equipment','DIFA equipment','Booked Sets','Sohar History','Repair History'}:
                        ws.row_dimensions[1].height=30
                        ws.column_dimensions['A'].width=65
                        ws.column_dimensions['B'].width=15
                        ws.column_dimensions['C'].width=15
                        ws.column_dimensions['D'].width=20
                        ws.column_dimensions['E'].width=11
                        ws.column_dimensions['F'].width=13
                        ws.column_dimensions['G'].width=11
                        ws.column_dimensions['H'].width=12
                        ws.freeze_panes = 'A2'
                        
                else:
                        ws.row_dimensions[1].height=30
                        ws.column_dimensions['A'].width=12.5
                        ws.column_dimensions['B'].width=65
                        ws.column_dimensions['C'].width=15
                        ws.column_dimensions['D'].width=20
                        ws.column_dimensions['E'].width=15
                        ws.column_dimensions['F'].width=11
                        ws.column_dimensions['G'].width=13
                        ws.column_dimensions['H'].width=12
                        ws.freeze_panes = 'A2'
                        

        wb.save(file+'.xlsx')
        
        
    def pn2desc_map(self):

        '''Contructs a dictionary with PN as keys and description as values. Will be used for  displaying min store in GUI.'''
        pnMap={}
        
        for k in self.DH:
                
                if self.DH[k][1] not in pnMap:
                        pnMap[self.DH[k][1]]=self.DH[k][0]
            
        for k in self.cables:
                
                if self.cables[k][1] not in pnMap:
                        pnMap[self.cables[k][1]]=self.cables[k][0]
        for k in self.consumables:
                if k not in pnMap:
                        pnMap[k]=self.consumables[k][0]
        
        return pnMap

    def validate(self,pn):
        '''For a given PN (string), the method compares the incoming qty for this PN with the qty of the available+used+booked+sent
        It attempts to validate whether the available + consumed is equal the incoming or not.'''
        history=0    
        try:
                av=self.av_pn_map()[pn]
        except:
                av=0
        try:
                inc=self.inc_pn_map()[pn]
        except:
                inc=0
        try:
                booked=self.booked_pn_map()[pn]
        except:
                booked=0
        try:
                used=self.used_pn_map()[pn]
        except:
                used=0
        for k in self.lekh:
                if pn in self.lekh[k]:
                        history+=1
        for k in self.nimr:
                if pn in self.nimr[k]:
                        history+=1
        
        
        if inc==av+used+booked+history:
##                print('Valid')
                return True,(inc,av,used,booked,history)
        else:
##                print('Invalid, '+str(inc-(av+used+booked+history))+' pieces missing.')
                return False,(inc,av,used,booked,history)


class WO(Store):##Need standalone function to get WO files from folder uses
                ##This class to create objects for them.
        
    def __init__(self,name,filename): ##DONE
        '''Creates WO object, uses Store init function but adds base attribute.
        WO status is unbooked by default. Should be either booked, unbooked
        or sent.'''
        self.file=filename
        self.name=name.upper()

        config=cp.ConfigParser()
        config.read(filename)

        self.DH=eval(config.get(self.name,"dh"))
        self.cables=eval(config.get(self.name,"cables"))
        self.consumables=eval(config.get(self.name,"consumables"))
        self.rogue=eval(config.get(self.name,"rogue"))
        self.base=config.get(self.name,"base")
        self.status=config.get(self.name,"status")
        self.sent=config.get(self.name,"sent")
        self.internal=eval(config.get(self.name,"internal"))
        
        
        
##        for k in self.cables:
##                if k in rogues:
##                        self.rogue=rogues[k].copy()
##                else:
##                        self.rogue=[]
    def set_base(self,base):
        self.base=base
        return
    def set_status(self,status,date=''):
        self.status=status
        return
    def set_name(self,name):           
        self.name=name.upper()
        return
    def get_base(self):
        return self.base

    def get_status(self):
        return self.status
    def save(self):
        config=cp.ConfigParser()
        config.read(self.file)
        config.set(self.get_name(),'name',self.name)
        config.set(self.get_name(),'dh',str(self.DH))
        config.set(self.get_name(),'cables',str(self.cables))
        config.set(self.get_name(),'consumables',str(self.consumables))
        config.set(self.get_name(),'rogue',str(self.rogue))
        config.set(self.get_name(),'base',str(self.base))
        config.set(self.get_name(),'status',str(self.status))
        config.set(self.get_name(),'sent',str(self.sent))
        config.set(self.get_name(),'internal',str(self.internal))
        f=open(self.file,'w')
        config.write(f)
        f.close()
        return
            
    def return_eqy(self,sohar,datein,used=False): ##DONE
        '''returns WO to sohar's store from sohar history. 'used' parameter
        specifies if the set
        was pulled or just returned as new. If returned set is new then add
        it to Store. If the returned set is used then add it to used store.
        datein must be specified to log in sohar history'''
        ##Need to think about marking it for DIFA.
        self.sent=''
        if not used:
                if self.get_status()=='Sent':
                    ##Mutate datein in WO dict
                        
                        for k in self.DH:
                                self.DH[k][5]=datein
                        for k in self.cables:
                                self.cables[k][5]=datein
                        sohar.add_WO(self)##sohar.add_WO(WO)
                        self.set_status('Unbooked') ##Mark WO as unbooked in excel and object
                        file_write(sohar,self,'shistory',datein,'')##file_write(WO,'history',datein,dateout='')
                        sohar.save()##sohar.save()
                        self.save()

                
                    
                else: ## if WO status is anything else raise exception.
##                        print('WO was not sent. Please check in booked sets.')
                        return 'WO not sent'

        elif used:
                if self.get_status()!='Sent':
                        return 'WO not sent'

                for k in self.DH:
                        self.DH[k][5]=datein
                        self.DH[k][6]='Used'
                for k in self.cables:
                        self.cables[k][5]=datein
                        self.cables[k][6]='Used'
                self.set_status('Used')
                for k in self.DH:
                        sohar.difa[k]=self.DH[k].copy()
                        sohar.difa[k].append(self.get_base())
                        sohar.difa[k].append(self.get_name())
                for k in self.cables:
                        sohar.difa[k]=self.cables[k].copy()
                        sohar.difa[k].append(self.get_base())
                        sohar.difa[k].append(self.get_name())

                file_write(sohar,self,'shistory',datein,dateout='')        
                sohar.save()        
                self.save()
                
    def is_avail(self,sohar): ##DONE
        ''' checks if WO is availbale in Sohar. Returns True if all items
            are available. Does not deduct items from Sohar.'''
        keys=list(self.DH.keys())+list(self.cables.keys())##Accumulate keys for DH & cables dicts
        states=[]               ##List to accumulate avail test for each item
        items=[]
        print(keys)##List to accumulate missing items
        for k in keys:
                if '/' in k:
                        continue
                s,q=sohar.find_item(k,'SN') ##For each key call find_item to look for it on sohar
                print(f'{k} is {s}.')
                if k in sohar.DH:
                        if s==True and q>=self.DH[k][4]:
                                states.append(s) ##If it is availbale in sohar and qty is adequate append true to states list
                        else:
                                states.append(False) ## else append false
                                items.append(self.DH[k][0])
                elif k in sohar.cables:                         ##If k not found in DH then lookm for it in cables
                        if s==True and q>=self.cables[k][4]:
                                states.append(s)
                        else:
                                states.append(False)
                                items.append(self.cables[k][0])
                print(states)
        ##same as above but for consumbales
        keys=list(self.consumables.keys())
        for k in keys:
                s,q=sohar.find_item(k,'PN')
                if s==True and q>=self.consumables[k][4]:
                        states.append(s)
                else:
                        states.append(False)
                        items.append(self.consumables[k][0])
        if set(states)=={True}: ##If all items are available thena all items in states is True
                                ## Then converting to a set should have 1 True.
                ##Return either true or false to reflect if WO is available or not.
                return True,[]
        else:
                return False,items #Return False and list of missing items
    def reass_set(self, old_name, new_name,status, sohar,base):
        
        '''Reassign WO if it is booked or sent.'''
            ##Check if WO is booked or sent
        
        config=cp.ConfigParser()
        config.read(self.file)
        config.add_section(new_name)
        config.remove_section(old_name)
        self.set_name(new_name)
        self.set_base(base)
        f=open(self.file,'w')
        config.write(f)
        f.close()
        self.save()
        if status=='Booked':
                for L in sohar.bookings:
                        if L[-1]==old_name:
                                L[-1]=new_name
                                L[-2]=base
                sohar.save()
                return
        elif status=='Sent':

                for row in sohar.shistory:
                        if row[7]==self.sent and row[-1]==old_name:
                                row[-1]=new_name
                                row[-2]=base
                sohar.save()
                return
                
                
            ##if booked, retrieve booked lists and change WO name in bookings
            ## if sent, retrieve sohar history and change well name in it.
    def book_set(self,sohar):##DONE
        '''Books the WO. Deducts from sohar and Moves it to Booked tab excel
        file. WO file should be marked as booked.
        Should change self.booking to True.'''
        state=sohar.sub_WO(self) ##Deduct WO from sohar store
        if state: ##If deduction was successful
                sohar.clear_ZQ() ##clear zeroes from sohar store
                #Loop over WO dict and add it to sohar bookings store
                for k in self.DH: 
                        a=self.DH[k].copy()
                        a.append(self.get_base())
                        a.append(self.name)
                        sohar.bookings.append(a)
                for k in self.cables:
                        a=self.cables[k].copy()
                        a.append(self.get_base())
                        a.append(self.name)
                        sohar.bookings.append(a)
                for k in self.consumables:
                        a=self.consumables[k].copy()
                        a.extend([None,None])
                        a.append(self.get_base())
                        a.append(self.name)
                        sohar.bookings.append(a)
                sohar.save()
        else:
                return 'Not Available'
        
        self.set_status('Booked') ##Mark WO as booked in excel and object
        self.save()
        print('wo saved')
        
    def unbook_set(self,sohar): ##DONE
        '''Removes WO from Booked tab in excel file to Sohar store.
        Changes self.booking to False.'''
        sohar.add_WO(self) ##Add WO back to sohar store
        book=[]
        for row in sohar.bookings: ##Loop over each row in booking dict
                if row[-1]!=self.get_name(): ## If last well name (last item in list) is not the same as well name 
                        book.append(row) ## Accumulate entries for booked eqy of other wells.
        sohar.bookings=book ##Move the remaining booked equipment to sohar.bookings
                
        sohar.save() ##save sohar store
        self.set_status('Unbooked') ##Mark WO as unbooked in excel and object
        self.save()

    def add_item(self,sohar, data,dateout=None): ##DONE, augmented after testing
        '''WO editing tool. Used to add item to WO. Checks to see if WO is unbooked or
        if data is either len 6 or 3 or if it is available in sohar store, if above consitions are met
        it deducts the item from sohar store and adds item to WO object and file. It then checks if WO is booked or sent
        If booked it adds the item to sohar bookings attribute and file. If WO is sent then it acts as if the item will be sent
        so it adds entry to sohar history using the dateout supplied in the input argument. User must specify this date or it will
        be left blank
        '''

        if len(data)==7:##Check if item is in store.
                s,q=sohar.find_item(data[3],'SN')
                
        elif len(data)==5:
                s,q=sohar.find_item(data[1],'PN')
        elif len(data)==6:
                s,q=sohar.find_item(data[5],'SN')
        else:
##                print('Incorrect data format.')
                return 'Incorrect data format.'
        
        if self.get_status()=='Unbooked':
                if s:
                        if len(data)==7:
                                self.DH[data[3]]=data ##Add item to WO
                        elif len(data)==5:
                                if data[1] in self.consumables: ##Check if item is already in WO
                                        self.consumables[data[1]][4]+=data[4] ## increment qty
                                else:
                                        self.consumables[data[1]]=data ##add new entry
                        elif len(data)==6:
                                a=cable_decode(data[-1],data[:-1])
                                for i in range(len(a)):
                                        if i==0:
                                                self.cables[data[-1]]=a[i]
                                        else:
                                                self.cables[data[-1]+'/'+str(i+1)]=a[i]

                        
                else:
                        return 'Item not in store.'
 
        
        elif self.get_status()=='Booked':##If it is Booked:
                ##Add item to bookings lists
                

                return 'set booked'

        elif self.get_status()=='Sent':##if it is sent:
                ##Add item to sohar history dict
                 if s:
                        if len(data)==7:
                                self.DH[data[3]]=data ##Add item to WO
                                sohar.DH[data[3]][4]-=data[4]
                                data2=data.copy()
                                data2.extend([dateout,self.get_base(),self.get_name()])
                                sohar.shistory.append(data2)
                        elif len(data)==5:
                                if q<data[4]:
                                        return 'Item not in store.'
                                sohar.consumables[data[1]][4]-=data[4]
                                if data[1] in self.consumables: ##Check if item is already in WO
                                        self.consumables[data[1]][4]+=data[4] ## increment qty      
                                else:
                                        self.consumables[data[1]]=data ##add new entry
                                data2=data.copy()
                                data2.extend(['','',dateout,self.get_base(),self.get_name()])
                                sohar.shistory.append(data2)        
                        elif len(data)==6:
                                a=cable_decode(data[-1],data[:-1])
                                for i in range(len(a)):
                                        self.cables[data[-1]+'/'+str(i+1)]=a[i]
                                del(sohar.cables[data[-1]])
                                for r in a:
                                        r.extend([dateout,self.get_base(),self.get_name()])
                                        sohar.shistory.append(r)

        sohar.clear_ZQ()
        sohar.save() ##save sohar store      
        self.save()
        return

    def del_item(self,sohar,data,datein=None): ##DONE, augmented after testing
        '''WO editing tool. Used to remove item from WO. It checks if WO is not unbooked or item is not
        in sohar already if it has a SN or if item is not in WO. If the above conditions are satisfied then
        it proceeds to deduct the item from WO object and file. Next it adds item back to sohar store object
        It then checks to see if WO is Booked then it removes this item from sohar bookings attribute
        If WO was sent it acts as if the item is returned to sohart and adds an entry in sohar history with
        datein supplied in input arguments. If user does not specify the date it will be left blank.'''
        TD={}
        if self.get_status()=='Unbooked':##Check the WO status. Return False if it is Unbooked.
                
                if len(data)==7 and data[0][:3]!='Cab':
                        del(self.DH[data[3]])
                
                elif len(data)==5:
                        if data[4]==self.consumables[data[1]][4]:
                                del(self.consumables[data[1]])
                        else:
                                self.consumables[data[1]][4]-=data[4]
                elif len(data)==7 and data[0][:3]=='Cab':
                        for k in self.cables:
                                if data[3] in k:
                                        pass
                                else:
                                        TD[k]=self.cables[k].copy()
                        self.cables=TD
                self.save()
                return
                
                
        if len(data)==7 and data[0][:3]!='Cab': ##Check if item is in WO.
                if data[3] not in self.DH:
                        return 'Item not in '
                
        elif len(data)==5:
                if data[1] not in self.consumables:
                        return 'Item not in '
        elif len(data)==7 and data[0][:3]=='Cab':
                if data[3] not in self.cables:
                        return 'Item not in '
                        
        ##deduct from WO object 
        
        if self.get_status()=='Booked':
                return 'set booked'
        if self.get_status=='Sent':
                return 'Item already sent.'

if __name__=='__main__':
        
        s=Store('sohar','config.ini')
        w=WO('TN3','config.ini')
##        w.unbook_set(s)
        s.Move_set(w,'')

        
        
