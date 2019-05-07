##Edit 1 in Git Hub
##Need to ditch the load INV button and add delete WO button.
from PyQt5 import QtCore, QtGui, QtWidgets,Qt
import lic
import Store_ManagerV09
import os
import sys
import openpyxl
import Main
import wo_creator
import repair_form
import validation
import WO_Edit
import send_wo
import return_wo
import Return_cable
import new_batch
import WO_reassign
import minStore_dlg
import minStore_dlg2
import minStore_dlg3
import reel_update
import time
import shutil
import wells
import clear_difa
import clear_difa2##cables difa
import datetime
import cable_scrap

class Dashboard(QtWidgets.QMainWindow,Main.Ui_MainWindow):
    def __init__(self,parent=None):
        QtWidgets.QMainWindow.__init__(self,parent)
        self.setupUi(self)
        
##        self.store_viewer.customContextMenuRequested.connect(self.contextMenuEvent)
        self.status=self.statusBar()
        self.status.showMessage("Load Workorder")
        self.load_inv()
        self.create_wo.clicked.connect(self.wo_creation) ##OK
        self.save_action.triggered.connect(self.save_progress) ##OK
        self.saveas_action.triggered.connect(self.backup_store) ##OK
        self.st_selector.currentTextChanged.connect(self.tabView) ##OK Action for selecting from drop down menu
        self.load_store.clicked.connect(self.foo) ##Action for clicking load_store Button
        self.open_store.triggered.connect(self.foo)
        self.load_wo.clicked.connect(self.wo_load) ## OK Action for clicking Load WO Button
        self.open_WO.triggered.connect(self.wo_load) ##OK
        self.search_store.returnPressed.connect(self.filter)
        self.repair.clicked.connect(self.start_repair) ##OK
        self.validate.clicked.connect(self.validate_item)##OK
        self.add_item.triggered.connect(self.WO_AddItem)##OK
        self.rem_item.triggered.connect(self.WO_SubItem)##OK
        self.book_WO.triggered.connect(self.WO_booking) ##OK
        self.Unbook_WO.triggered.connect(self.WO_unbooking)##OK
        self.send_WO.triggered.connect(self.WO_send) ##OK
        self.return_WO.triggered.connect(self.WO_return) ##OK
        self.batch_action.triggered.connect(self.new_batch) ##OK
        self.reass_WO.triggered.connect(self.WO_reassign)##OK
        self.minstore_action.triggered.connect(self.MinStoreDlg)  ##OK
        self.check_minstore.triggered.connect(self.minstore_check)  ##OK but found bug when non existent PN is inputted
        self.reel_edit.clicked.connect(self.Edit_reel) ## OK
        self.exp_rep.clicked.connect(self.generate_rep) ##OK
        self.export_action.triggered.connect(self.generate_rep) ##OK
        self.update_orders.triggered.connect(self.orders_update)##OK
        self.difa_clear.triggered.connect(self.clearDIFA)##OK
        self.return_cable.triggered.connect(self.cable_return)
        self.actionClear_DIFA_cables.triggered.connect(self.clearDIFA_cables)
        self.scrap.triggered.connect(self.scrap_cable)
        
####    def contextMenuEvent(self, event):
####        menu = QtWidgets.QMenu(self)
####        add_wo=menu.addAction("Add to WO",self.add2WO)
####        action = menu.exec_(QtGui.QCursor.pos())
##    def add2WO(self):
##        if self.WO.get_status()=='Unbooked':
##            pn=''
##            i=self.store_viewer.currentRow()
##            sn=self.store_viewer.item(i,3)
##            sn=sn.text()
##            data=self.sohar.DH[sn].copy()
##            if sn!='None':
##                if sn in self.sohar.DH:
##                    data=self.sohar.DH[sn].copy()
##                elif sn in self.sohar.cables:
##                    data=self.sohar.cables[sn].copy()
##                self.WO.add_item(self.sohar,data)
##            else:
##                pn=self.store_viewer.item(i,1)
##                pn=pn.text()
##                data=self.sohar.consumables[pn].copy()
##                self.WO.add_item(self.sohar,data)
##        else:
##            msg=QtWidgets.QMessageBox()
##            msg.setIcon(QtWidgets.QMessageBox.Critical)
##            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
##            msg.setWindowTitle("Error")
##            msg.setText("You cannot use this method if wororder is already sent or booked.")
##            msg.exec_()

    def scrap_cable(self):
        form=scrapping()
        form.exec_()
        if form.result()==1:
            a=form.reelno.text()
            b=int(form.qty.text())
            c=form.armor.currentText()
            d=form.date.text()
            flag=self.sohar.cable_scrap(a,b,c,d)
            if flag==True:
                msg=QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Information)
                msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msg.setWindowTitle("Success")
                msg.setText("Cable added to scrap.")
                msg.exec_()
            else:
                msg=QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Warning)
                msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msg.setWindowTitle("Warning")
                msg.setText("Quantity on Reel is insufficient. Please check again.")
                msg.exec_()
            self.tabView()
        else:
            pass

    def clearDIFA_cables(self):
        form=clearing_difa_cables()
        form.exec_()
        if form.result()==1:
            if form.sn_filter.text()=='':
                pass
            else:
                sn=form.sn_filter.text()
                self.sohar.used_cables[sn]=self.sohar.difa_cables[sn].copy()
                del(self.sohar.difa_cables[sn])
                self.tabView()
        else:
            pass
    def cable_return(self):##mark2
        form=Returning_cable(self.sohar)
        form.exec_()
        if form.result()==1:
            sn=''
            date=form.return_date.text()
            Base=form.base_select.currentText()
            data=[]
            
            for i in range(form.return_view.rowCount()):
                sn=form.return_view.item(i,0).text()
                
                data=[int(form.return_view.item(i,1).text()),int(form.return_view.item(i,2).text()),
                int(form.return_view.item(i,3).text()),int(form.return_view.item(i,4).text()),
                form.return_view.item(i,5).text()]
                print(data)
                well=form.return_view.item(i,6).text()
                enc={sn:data}
                print(enc)
                cable=Store_ManagerV09.cable_decode(sn,data)
                print(cable)
               
                print('Im gone.')
                self.sohar.return_cable(cable,enc,date,Base,well)
                print('Never Know what hit em!')
            self.tabView()       
            
    def clearDIFA(self):
        difaForm=Clearing_difa(self.sohar)
        difaForm.exec_()
        if difaForm.result()==1:
            sn=[]
            for i in range(difaForm.cleared_view.rowCount()):
                sn.append(difaForm.cleared_view.item(i,1).text())
            for s in sn:
                self.sohar.clear_difa(s)
            self.tabView()
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Success")
            msg.setText("Equipment cleared from DIFA.")
            msg.exec_()
        else:
            pass
        
    def orders_update(self):
        filename=QtWidgets.QFileDialog.getOpenFileName(QtWidgets.QMainWindow(),'Open File') ##Open file dialog to choose sohar store file
        if filename[0]!='':
            self.sohar.update_orders(filename[0])
            self.tabView()
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Success")
            msg.setText("Orders updated successfully.")
            msg.exec_()
    def wo_load(self):

        wellsList=wells_list(self.sohar.get_wos())
        wellsList.exec_()
        
        if wellsList.result()==1:
            name=wellsList.well_list.currentItem().text()
            self.WO=Store_ManagerV09.WO(str(name),self.sohar.file)
            self.tabView()
            
    def foo(self):
        term=self.search_store.text()
        for i in range(self.store_viewer.rowCount()):
            match=False
            item=self.store_viewer.item(i,9)
            if term in item.text():
                match=True
                
            self.store_viewer.setRowHidden(i,not match)
        
    def generate_rep(self):
        try:
            filename=QtWidgets.QFileDialog.getSaveFileName(QtWidgets.QMainWindow(),'Open File', ##Open file dialog to choose sohar store file
                                                           os.getenv('C:\\Users\\HP\\AppData\\Local\\Programs\\Python\\Python36-32\\Python exercises'))
            if filename[0].endswith('.xlsx'):
                f=filename[0].rstrip('.xlsx')
            else:
                f=filename[0]
            if filename[0]=='':
                return
            
            self.sohar.generate_report(f)
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Success")
            msg.setText("Report generated successfully.")
            msg.exec_()
        except:
            pass

    def Edit_reel(self):
        cabredit=reelUpdateDlg()
        cabredit.exec_()
        if cabredit.result()==1:
            a=self.sohar.update_reel(cabredit.SN1.text(),cabredit.SN2.text(),int(cabredit.length.value()),cabredit.arm_cond.currentText(),cabredit.well.text(),
                                     cabredit.date.text())
            if a==False:
                msg=QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Warning)
                msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msg.setWindowTitle("Warning")
                msg.setText("Insufficient Cable length.")
                msg.exec_()
                
            self.load_histories()
            self.tabView()
            
    def minstore_check(self):
        minstore=MinStore_Dlg2()
        minstore.exec_()
        
        if minstore.result()==1:
            pn_map=self.sohar.av_pn_map()
            if minstore.checkBox.isChecked():
                accum=[]
                for pn,qty in pn_map.items():
                    if pn in self.sohar.minStore_map:
                        if (qty-self.sohar.minStore_map[pn])>0 and (qty-self.sohar.minStore_map[pn])<=5:
                            accum.append((pn,qty,'Close to minimum'))
                        elif (qty-self.sohar.minStore_map[pn])==0 or (qty-self.sohar.minStore_map[pn])<0:
                            accum.append((pn,qty,'Below/at minimum'))
                minstoreView=MinStore_Dlg3()
                minstoreView.table.setHorizontalHeaderLabels(['Part Number','QTY','Status'])
                minstoreView.table.setSortingEnabled(False)
                for i in range(0,len(accum)):
                    minstoreView.table.insertRow(i)
                    minstoreView.table.setItem(i,0,QtWidgets.QTableWidgetItem(accum[i][0]))
                    minstoreView.table.setItem(i,1,QtWidgets.QTableWidgetItem(str(accum[i][1])))
                    minstoreView.table.setItem(i,2,QtWidgets.QTableWidgetItem(accum[i][2]))
                    minstoreView.table.resizeColumnToContents(0)
                    minstoreView.table.resizeColumnToContents(1)
                    minstoreView.table.resizeColumnToContents(2)
                minstoreView.table.setSortingEnabled(True)
                minstoreView.exec_()
                
                    
        
            else:
                pn=minstore.lineEdit.text()
                qty1=pn_map[pn]
                qty2=self.sohar.minStore_map[pn]
                qty=qty1-qty2
                if qty>0 and qty<=5:
                    ##raise warning
                    msg=QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Warning)
                    msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msg.setWindowTitle("Warning")
                    msg.setText("Min store quantity is "+str(qty2)+" and actual quantity is "+str(qty1)+". It is close to min quantity. Please monitor closely.")
                    msg.exec_()
                elif qty==0 or qty<0:
                    ##Raise critical alarm
                    msg=QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)
                    msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msg.setWindowTitle("Error")
                    msg.setText("Min store quantity is "+str(qty2)+" and actual quantity is "+str(qty1)+". Please order as soon as possible.")
                    msg.exec_()
                elif qty>5:
                    ## Satisfactory levels
                    msg=QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Information)
                    msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msg.setWindowTitle("Min Store Check")
                    msg.setText("Min store quantity is "+str(qty2)+" and actual quantity is "+str(qty1)+". Store levels are good.")
                    msg.exec_()
            
        else:
            del(minstore)
   

    def MinStoreDlg(self):
        minstore=MinStore_Dlg(self)
        
        try:
            k=list(self.sohar.minStore_map.keys())
            for i in range(0,len(k)):
                minstore.pn_table.insertRow(i)
                minstore.pn_table.setItem(i,0,QtWidgets.QTableWidgetItem(k[i]))
                minstore.pn_table.setItem(i,1,QtWidgets.QTableWidgetItem(str(self.sohar.minStore_map[k[i]])))
                minstore.pn_table.resizeColumnToContents(0)
                minstore.pn_table.resizeColumnToContents(1)
            minstore.exec_()
            if minstore.result()==1:

                pn=[]
                qty=[]
                for i in range(0,minstore.pn_table.rowCount()):
                    pn.append(minstore.pn_table.item(i,0).text())
                    qty.append(minstore.pn_table.item(i,1).text())

                self.sohar.minStore_map={}
                for p,q in zip(pn,qty):
                    self.sohar.min_store(str(p),int(q))
        except:
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Warning")
            msg.setText("Please load Sohar inventory file.")
            msg.exec_()
            minstore.close()
 

    def WO_reassign(self):
        reassign=Reassignment()
        try:
            reassign.fromWell.setText(self.WO.get_name())
            reassign.exec_()
        except:
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Critical)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Error")
            msg.setText("Please load workorder file first.")
            msg.exec_()
        if reassign.result()==1:
            
            old_name=self.WO.get_name()
            new_name=reassign.toWell.text()
            self.WO.reass_set(old_name.upper(),new_name.upper(),self.WO.get_status(),self.sohar,reassign.base.currentText())
            self.load_histories()
            self.tabView()
            

    def new_batch(self):
        state=False
        while state==False:
            batch=Batch_dlg()
            batch.exec_()
            if batch.result()==1 and batch.batchFile.text()!='':
                state=True
                try:
                    check,inv=self.sohar.new_batch(batch.file,batch.batchDate.text())
                    self.load_histories()
                    self.tabView()
                    if len(check)==0:
                        msg=QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Information)
                        msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msg.setWindowTitle("Success")
                        msg.setText("New Equipment Batch added to Sohar Store.")
                        msg.exec_()
                    else:
                        msg=QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Warning)
                        msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msg.setWindowTitle("Warning")
                        msg.setText("Some items were already found in Sohar Store.\nWould you like to see a list of SN?")
                        msg.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                        res=msg.exec_()
                        if res ==QtWidgets.QMessageBox.Yes:
                            msg2=QtWidgets.QMessageBox()
                            msg2.setIcon(QtWidgets.QMessageBox.Information)
                            msg2.setWindowIcon(QtGui.QIcon('icon2.ico'))
                            msg2.setWindowTitle("List of Items")   
                            msg2.setText(str(check))
                            msg2.exec_()
                        else:
                            pass
                    
                except:
                    msg=QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Critical)
                    msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msg.setWindowTitle("Error")
                    msg.setText("Please load Sohar Inventory file.")
                    msg.exec_()
                    
            elif batch.result()==1 and batch.batchFile.text()=='':
                msg=QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Warning)
                msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msg.setWindowTitle("Warning")
                msg.setText("Please select Equipment Batch file.")
                msg.exec_()
            else:
                state=True

    def WO_return(self):##mark2

        returnForm=Returning_set(sohar=self.sohar)
        returnForm.exec_()
        
        if returnForm.result()==1:
            sn=[]
            qty=[]
            condition=[]
            well=[]
            date=returnForm.return_date.text()
            Base=returnForm.base_select.currentText()
            data=[]
            for i in range(returnForm.return_view.rowCount()):
                sn.append(returnForm.return_view.item(i,1).text())
                qty.append(returnForm.return_view.item(i,2).text())
                condition.append(returnForm.return_view.item(i,3).text().capitalize())
                well.append(returnForm.return_view.item(i,4).text().upper())
            
            for s,q,c,w in zip(sn,qty,condition,well):
                data.append((s,q,c,w))

            self.sohar.return_item(data,date,Base)
            self.tabView()               

    def generate_TN(self,date,base,well,dh={},cables={},cons={},job=''):##findme2
        filename=QtWidgets.QFileDialog.getSaveFileName(QtWidgets.QMainWindow(),'Save Transfer Note')
        f=filename[0]
        if f.endswith('.xlsx'):
            shutil.copyfile('tn.xlsx',f)
        else:
            f=filename[0]+'.xlsx'
            shutil.copyfile('tn.xlsx',f)
            
        wb=openpyxl.load_workbook(f)
        ws=wb['tn']
        row1=32
        row2=27
        dash=openpyxl.styles.borders.Side(style='dashed')
        for i,k in zip(range(row1,row1+len(cons)),cons):
            ws.insert_rows(i)
            ws['C'+str(i)]=cons[k][2]
            ws['D'+str(i)]=cons[k][1]
            ws['E'+str(i)]=''
            ws['F'+str(i)]=cons[k][0]
            ws['H'+str(i)]=cons[k][4]
            ws['I'+str(i)]=''
            for c in ws[str(i)]:
                c.font=openpyxl.styles.Font(name='Calibri',sz=13,color='003366')
                c.border=openpyxl.styles.Border(left=dash,top=dash,bottom=dash,right=dash)
        for i,k in zip(range(row2,row2+len(cables)),cables):
            ws.insert_rows(i)
            ws['C'+str(i)]=cables[k][2]
            ws['D'+str(i)]=cables[k][1]
            ws['E'+str(i)]=cables[k][3]
            ws['F'+str(i)]=cables[k][0]
            ws['H'+str(i)]=cables[k][4]
            ws['I'+str(i)]=cables[k][6]
            for c in ws[str(i)]:
                c.font=openpyxl.styles.Font(name='Calibri',sz=13,color='003366')
                c.border=openpyxl.styles.Border(left=dash,top=dash,bottom=dash,right=dash)
        for i,k in zip(range(row2,row2+len(dh)),dh):
            ws.insert_rows(i)
            ws['C'+str(i)]=dh[k][2]
            ws['D'+str(i)]=dh[k][1]
            ws['E'+str(i)]=dh[k][3]
            ws['F'+str(i)]=dh[k][0]
            ws['H'+str(i)]=dh[k][4]
            ws['I'+str(i)]=dh[k][6]
            for c in ws[str(i)]:
                c.font=openpyxl.styles.Font(name='Calibri',sz=13,color='003366')
                c.border=openpyxl.styles.Border(left=dash,top=dash,bottom=dash,right=dash)
        ws['E17']=date
        ws['E18']=well
        ws['E20']=base
        if job=='external':
            ws['E16']='S'+str(datetime.datetime.today().year)[-2:]+str(self.sohar.jobCount)
        elif job=='internal':
            ws['E16']='JO'+str(datetime.datetime.today().year)[-2:]+str(self.sohar.jobCount)
        else:
            ws['E16']=job
        self.sohar.jobCount+=1
        ws.merge_cells('B'+str(28+len(dh)+len(cables)+1)+':J'+str(28+len(dh)+len(cables)+1))
        ws['B'+str(28+len(dh)+len(cables)+1)].border=openpyxl.styles.Border(left=dash,top=dash,bottom=dash,right=dash)
        wb.save(f)
        
        
        
    
    def WO_send(self):
        state=False
        while state==False:
            date=Sending_set()
            date.exec_()
            if date.result()==1:
                if date.send_date.text()=='':
                    state=False
                    msg=QtWidgets.QMessageBox()
                    msg.setIcon(QtWidgets.QMessageBox.Warning)
                    msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msg.setWindowTitle("Warning")
                    msg.setText("Please enter date of sending.")
                    msg.exec_()
                else:
                    try:
                        flag=self.sohar.Move_set(self.WO,date.send_date.text())
                        if flag=='Already sent':
                            msg=QtWidgets.QMessageBox()
                            msg.setIcon(QtWidgets.QMessageBox.Critical)
                            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                            msg.setWindowTitle("Error")
                            msg.setText(self.WO.get_name()+" already sent to field.")
                            msg.exec_()
                        else:
                            msg=QtWidgets.QMessageBox()
                            msg.setIcon(QtWidgets.QMessageBox.Information)
                            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                            msg.setWindowTitle("Success")
                            msg.setText(self.WO.get_name()+" was sent successfully.")
                            msg.exec_()
                            self.load_histories()
                            self.tabView()
                            self.generate_TN(date=date.send_date.text(),base=self.WO.get_base(),well=self.WO.get_name(),dh=self.WO.DH,cables=self.WO.cables,cons=self.WO.consumables,job='external')
                        state=True
                    except:
                        msg=QtWidgets.QMessageBox()
                        msg.setIcon(QtWidgets.QMessageBox.Critical)
                        msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msg.setWindowTitle("Error")
                        msg.setText("Please load inventory or WO files.")
                        msg.exec_()
                        
                        
            else:
                state=True
                
        
    
    def WO_unbooking(self):
        try:
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Unbook Set")
            msg.setText("Do you want to unbook "+self.WO.get_name()+" set?")
            msg.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            r=msg.exec_()
        except:
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Warning")
            msg.setText("Please load Workorder file.")
            msg.exec_()
            return
            
        if r==QtWidgets.QMessageBox.Yes:
            self.WO.unbook_set(self.sohar)
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Success")
            msg.setText(self.WO.get_name()+" set is unbooked.")
            msg.exec_()
            self.tabView()
        elif r==QtWidgets.QMessageBox.No:
            pass
        

        

    def WO_booking(self):
        try:
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Book Set")
            msg.setText("Do you want to book "+self.WO.get_name()+" set?")
            msg.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            r=msg.exec_()
        except:
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Warning)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Warning")
            msg.setText("Please load Workorder file.")
            msg.exec_()
            return 
            
        if r==QtWidgets.QMessageBox.Yes:
            flag=self.WO.book_set(self.sohar)
            print('booked')
            if flag=='Not Available':
                msg=QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Critical)
                msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msg.setWindowTitle("Error")
                msg.setText("One or more items in "+self.WO.get_name()+" are not available in Sohar store.")
                msg.exec_()
                return
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Success")
            msg.setText(self.WO.get_name()+" set booked.")
            msg.exec_()
            self.tabView()
        elif r==QtWidgets.QMessageBox.No:
            pass

    def WO_SubItem(self):
        subtraction=ItemSubDlg()
        subtraction.exec_()
        if subtraction.result()==1:
            
            if subtraction.sn_option.isChecked():
                if subtraction.id_input.text() in self.WO.DH:
                    data=self.WO.DH[subtraction.id_input.text()].copy()
                    data[4]=subtraction.spinBox.value()
                    flag=self.WO.del_item(self.sohar,data,subtraction.date_input.text())
                    self.load_histories()
                    self.tabView()
                    if flag=='Invalid data.':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Incorrect data format")
                        msgBox.exec_()
                    elif flag=='Item not in ':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Item was not found in workorder "+self.WO.get_name())
                        msgBox.exec_()
                    elif flag=='Item already in Sohar':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Item is already in Sohar store. It was removed from workorder before.")
                        msgBox.exec_()
                            
                elif subtraction.id_input.text() in self.WO.cables:
                    data=self.WO.cables[subtraction.id_input.text()].copy()
                    data[4]=subtraction.spinBox.value()
                    
                    
                    flag=self.WO.del_item(self.sohar,data,subtraction.date_input.text())
                    self.load_histories()
                    self.tabView()
                    if flag=='Invalid data.':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Incorrect data format")
                        msgBox.exec_()
                    elif flag=='Item not in ':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Item was not found in workorder "+self.WO.get_name())
                        msgBox.exec_()
                    elif flag=='Item already in Sohar':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Item is already in Sohar store. It was removed from workorder before.")
                        msgBox.exec_()
                else:
                    msgBox=QtWidgets.QMessageBox()
                    msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                    msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msgBox.setWindowTitle("Error")
                    msgBox.setText("Item not in Workorder.")
                    msgBox.exec_()
                    
            elif subtraction.pn_option.isChecked():
                data=self.WO.consumables[subtraction.id_input.text()].copy()
                data[4]=subtraction.spinBox.value()
                
                
                flag=self.WO.del_item(self.sohar,data,subtraction.date_input.text())
                self.load_histories()
                self.tabView()
                if flag=='Invalid data.':
                    msgBox=QtWidgets.QMessageBox()
                    msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                    msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msgBox.setWindowTitle("Error")
                    msgBox.setText("Incorrect data format")
                    msgBox.exec_()
                elif flag=='Item not in ':
                    msgBox=QtWidgets.QMessageBox()
                    msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                    msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msgBox.setWindowTitle("Error")
                    msgBox.setText("Item was not found in workorder "+self.WO.get_name())
                    msgBox.exec_()
                elif flag=='Item already in Sohar':
                    msgBox=QtWidgets.QMessageBox()
                    msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                    msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msgBox.setWindowTitle("Error")
                    msgBox.setText("Item is already in Sohar store. It was removed from workorder before.")
                    msgBox.exec_()
            else:
                msgBox=QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msgBox.setWindowTitle("Warning")
                msgBox.setText("Please select ID type (SN/PN).")
                msgBox.exec_()
            
        else:
            del(subtraction)

    def WO_AddItem(self):
        addition=ItemAddDlg()
        addition.exec_()
        if addition.result()==1:
            
            if addition.sn_option.isChecked():
                if addition.id_input.text() in self.sohar.DH:
                    data=self.sohar.DH[addition.id_input.text()].copy()
                    data[4]=addition.spinBox.value()
                    flag=self.WO.add_item(self.sohar,data,addition.date_input.text())
                    self.load_histories()
                    self.tabView()
                    if flag=='Incorrect data format.':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Incorrect data format")
                        msgBox.exec_()
                    elif flag=='Item not in store.':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Item is not available in Sohar Store.")
                        msgBox.exec_()
                            
                elif addition.id_input.text() in self.sohar.cables:
                    data=self.sohar.cables[addition.id_input.text()].copy()
                    data[4]=addition.spinBox.value()
                    flag=self.WO.add_item(self.sohar,data,addition.date_input.text())
                    self.load_histories()
                    self.tabView()
                    if flag=='Incorrect data format.':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Incorrect data format")
                        msgBox.exec_()
                    elif flag=='Item not in store.':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Item is not available in Sohar Store.")
                        msgBox.exec_()
                else:
                    msgBox=QtWidgets.QMessageBox()
                    msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                    msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msgBox.setWindowTitle("Error")
                    msgBox.setText("Item is not available in Sohar Store.")
                    msgBox.exec_()
                    
            elif addition.pn_option.isChecked():
                data=self.sohar.consumables[addition.id_input.text()].copy()
                data[4]=addition.spinBox.value()                
                flag=self.WO.add_item(self.sohar,data,addition.date_input.text())
                self.load_histories()
                self.tabView()
                if flag=='Incorrect data format.':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Incorrect data format")
                        msgBox.exec_()
                elif flag=='Item not in store.':
                        msgBox=QtWidgets.QMessageBox()
                        msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                        msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                        msgBox.setWindowTitle("Error")
                        msgBox.setText("Item is not available in Sohar Store.")
                        msgBox.exec_()
            else:
                
                msgBox=QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msgBox.setWindowTitle("Warning")
                msgBox.setText("Please select ID type (SN/PN).")
                msgBox.exec_()
            
        else:
            del(addition)
        
        
    def validate_item(self):
        valid=Validation(self)
        valid.exec_()
        if valid.result()==1:
            a,b=self.sohar.validate(valid.pn_input.text())
            if a==True:
                msgBox=QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Information)
                msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msgBox.setWindowTitle("Success")
                msgBox.setText(f"Item with PN: {valid.pn_input.text()} has a valid count. Incoming: {b[0]}, Available: {b[1]}, Booked: {b[3]}, Sent: {b[4]} and Used: {b[2]}.")
                msgBox.exec_()
            else:
                msgBox=QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msgBox.setWindowTitle("Error")
                msgBox.setText(f"Item with PN: {valid.pn_input.text()} does not have a valid count. Incoming: {b[0]}, Available: {b[1]}, Booked: {b[3]}, Sent: {b[4]} and Used: {b[2]}.")
                msgBox.exec_()
        else:
            del(valid)
        
    
    def start_repair(self):
        
        repTool=Repair_dlg(self)
        repTool.exec_()
        

        try:
            if len(repTool.sp_list)==0:
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msgBox.setWindowTitle("Warning")
                msgBox.setText("No spare parts selected for the repair.\n Are you sure you want to proceed?")
                msgBox.setStandardButtons(QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
                res=msgBox.exec_()
                if res==QtWidgets.QMessageBox.Yes:
                    pass
                else:
                    msgBox = QtWidgets.QMessageBox()
                    msgBox.setIcon(QtWidgets.QMessageBox.Warning)
                    msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                    msgBox.setWindowTitle("Warning")
                    msgBox.setText("Item repair is suspended")
                    msgBox.exec_()
                    return
            a,b=self.sohar.repair(str(repTool.sn_input.text()),repTool.sp_list,str(repTool.rep_date.text()))
            if a=='Insufficient':
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msgBox.setWindowTitle("Error")
                msgBox.setText("Insufficient quantity from spare part with PN: "+b)
                msgBox.exec_()
            elif a=='Itemnot':
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Critical)
                msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msgBox.setWindowTitle("Error")
                msgBox.setText("Item not found in used store.")
                msgBox.exec_()
            else:
                msgBox = QtWidgets.QMessageBox()
                msgBox.setIcon(QtWidgets.QMessageBox.Information)
                msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msgBox.setWindowTitle("Success")
                msgBox.setText("Item repaired successfully.")
                msgBox.exec_()
                self.load_histories()
                self.tabView()
        except:
            pass
    def wo_deletion(self):
        pass
    def wo_creation(self):
        woTool=WO_creator(self)
        woTool.setModal(False)
        woTool.exec_()
        

        
        try:##findme4
            inter=self.sohar.create_wo(woTool.well_name,woTool.base_name,woTool.sn_list,woTool.pn_list)
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Information)
            msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msgBox.setWindowTitle("Success")
            msgBox.setText("Workorder created successfully.")
            msgBox.exec_()
            for i in inter:
                self.generate_TN(date=datetime.datetime.today().strftime('%d-%b-%Y'),base=woTool.base_name,well=woTool.well_name,dh={},cables={},cons=i,job='internal')
            
        except:
            pass

    def load_histories(self): 
        pass        
            
    def load_inv(self): 
        
        try:
            self.sohar=Store_ManagerV09.Store('sohar','config.ini') ## Create sohar object from selected file
            self.tabView()
            self.load_histories()

        except: ## If no file is selected from dialog (user closed dialog), then do nothing.
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Critical)
            msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msgBox.setWindowTitle("Error")
            msgBox.setText("Configuration file missing!")
            msgBox.exec_()
        
    
        
    def backup_store(self):
        filename=QtWidgets.QFileDialog.getSaveFileName(QtWidgets.QMainWindow(),'Save File')
        try:
            f=filename[0].rstrip('.xlsx')
            shutil.copyfile(self.sohar.file,f+'.ini')
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Success")
            msg.setText("Progress Saved.")
            msg.exec_()
        except:
            pass
    def save_progress(self):
        try:
            self.sohar.save()
            try:
                self.WO.save()
            except:
                pass
            msg=QtWidgets.QMessageBox()
            msg.setIcon(QtWidgets.QMessageBox.Information)
            msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msg.setWindowTitle("Success")
            msg.setText("Progress Saved.")
            msg.exec_()
        except:
            pass
        
    def search_by(self,IDtype,ID,view):
        VD1={'Sohar Downhole equipment':self.sohar.DH,'Sohar Cables':self.sohar.cables,
            'Sohar Consumables':self.sohar.consumables,
            'Sohar Used equipment':self.sohar.used, 'Sohar DIFA equipment':self.sohar.difa}
        
        VD2={'Sohar History':self.sohar.shistory,'Repair History':self.sohar.rhistory,'Cable History':self.sohar.chistory,
            'Booked Sets':self.sohar.bookings}

        IDD={'Description':0,'Part Number':1,'1C Code':2, 'Serial Number':3}
        data=[]
        if view in VD1:
            for k in VD1[view]:
                if ID.lower() in str(VD1[view][k][IDD[IDtype]]).lower():
                    data.append(VD1[view][k].copy())
                else:
                    pass
        elif view in VD2:
            for L in VD2[view]:
                if ID.lower() in str(L[IDD[IDtype]]).lower():
                    data.append(L.copy())
                else:
                    pass
        elif view=='Minimum Store Levels':
            if IDtype=='1C Code' or IDtype=='Serial Number':
                msg=QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Warning)
                msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msg.setWindowTitle("Warning")
                msg.setText("You cannot search by 1C or SN in Min Store view.")
                msg.exec_()
                return []
            for k in self.sohar.minStore_map :
                    if k in self.sohar.orders:
                        a=self.sohar.orders[k][0]
                        b=self.sohar.orders[k][1]
                        c=self.sohar.orders[k][2]
                    else:
                        a=None
                        b=None
                        c=None

                    L=[self.sohar.pn2desc_map()[k],k,self.sohar.av_pn_map()[k],self.sohar.minStore_map[k],a,b,c]
                    if ID.lower() in str(L[IDD[IDtype]]).lower():
                        data.append(L)
    
        return data

        
        
    def filter(self):
        a=self.filter_mode.currentText()
        b=self.search_store.text()
        c=self.st_selector.currentText()

        if self.search_store.text()=='':
            self.tabView()
            return
        data=self.search_by(a,b,c)
        tabchk1={'Sohar Downhole equipment','Sohar Cables','Spare Parts'}
        tabchk2={'Sohar Used equipment','Booked Sets','Sohar DIFA equipment'}
        
        self.store_viewer.clear()
        self.store_viewer.setRowCount(len(data))
        if self.st_selector.currentText() in tabchk1:
            self.store_viewer.setHorizontalHeaderLabels(['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition'])
        elif self.st_selector.currentText() in tabchk2:
            self.store_viewer.setHorizontalHeaderLabels(['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition','Base','Well'])
        elif self.st_selector.currentText() == 'Sohar Consumables':
            self.store_viewer.setHorizontalHeaderLabels(['Description','Part Number','1C Code','Serial Number','Quantity'])
        elif self.st_selector.currentText() == 'Sohar History':
            self.store_viewer.setHorizontalHeaderLabels(['Description','Part Number','1C Code','Serial Number','Quantity','Date In',
                                                         'Condition','Date Out','Base','Well'])
        elif self.st_selector.currentText() =='Repair History':
            self.store_viewer.setHorizontalHeaderLabels(['Description','Part Number','1C Code','Serial Number','Quantity','Date In',
                                                         'Condition','Repair date'])
        elif self.st_selector.currentText() =='Cable History':
            self.store_viewer.setHorizontalHeaderLabels(['Description','Part Number','1C Code','Serial Number','Old Quantity','New Quantity',
                                                         'Delta','From Reel','To Reel'])
        elif self.st_selector.currentText() =='Minimum Store Levels':
            self.store_viewer.setHorizontalHeaderLabels(['Description','Part Number','Av Quantity','Min Store required','Intransit order','Intransit QTY','ETA'])
        j=0
        if len(data)==0:
            return
            
        for r in data:
            for i in range(0,len(r)):
                item=QtWidgets.QTableWidgetItem(str(r[i]))
                if i!=0:
                    item.setTextAlignment(128 | 4)
                self.store_viewer.setItem(j,i, item)
            j+=1
        for i in range(0,len(data[0])):
            self.store_viewer.resizeColumnToContents(i)
    
        
        
        
                       
    def tabView(self): ## Method to display the sohar attributes in the table widget
        self.store_viewer.clear() ## Start by clearing the table completely
        self.store_viewer.setSortingEnabled(False)
        ## Create Headers for each table view. Depending on the combobox selection the correct table will be loaded.
        header=['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition']

        try:
            if self.st_selector.currentText()=='Sohar Downhole equipment': ## Sohar Downhole is selected
                self.store_viewer.setRowCount(len(self.sohar.DH))
                self.store_viewer.setColumnCount(len(header))
                for i in range(0,len(header)): ##Load correct header and copy it to table
                    self.store_viewer. setHorizontalHeaderItem(i,QtWidgets.QTableWidgetItem(header[i]))
                j=0
                for r in self.sohar.DH: ## Loop over sohar.DH dict keys
                    
                    for i in range(0,len(self.sohar.DH[r])): ## loop over the list value of the given key
                        item=QtWidgets.QTableWidgetItem(str(self.sohar.DH[r][i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        else:
                            pass
                        self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i, QtWidgets.QTableWidgetItem(str(self.sohar.DH[r][i]))) ## Copy list items to table row
                    j+=1
                
                self.store_viewer.resizeColumnToContents(0)
                self.store_viewer.resizeColumnToContents(1)
                self.store_viewer.resizeColumnToContents(2)

                        
            elif self.st_selector.currentText()=='Sohar Cables': ## Sohar cables is selected
                header=['Reel no','Galv New','Galv Used','SS New','SS Used','Size']
                self.store_viewer.setRowCount(len(self.sohar.cables))
                self.store_viewer.setColumnCount(len(header))
                
                for i in range(0,len(header)): ##Load correct header and copy it to table
                    self.store_viewer. setHorizontalHeaderItem(i,QtWidgets.QTableWidgetItem(header[i]))
                j=0
                for r in self.sohar.cables:
                    item=QtWidgets.QTableWidgetItem(str(r))
                    item.setTextAlignment(128 | 4)
                    self.store_viewer.setItem(j,0, item)
                    j+=1
                    
                j=0
                for r in self.sohar.cables: ## Loop over sohar.cables dict keys
                    for i in range(0,len(self.sohar.cables[r])): ## loop over the list value of the given key
                        item=QtWidgets.QTableWidgetItem(str(self.sohar.cables[r][i]))
                        item.setTextAlignment(128 | 4)
                        self.store_viewer.setItem(j,i+1, item)
                    j+=1

                for i in range(len(header)):
                    self.store_viewer.resizeColumnToContents(i)
                    
            elif self.st_selector.currentText()=='Sohar DIFA Cables': ## Sohar cables is selected
                header=['Reel no','Galv New','Galv Used','SS New','SS Used','Size','Base','Well']
                self.store_viewer.setRowCount(len(self.sohar.difa_cables))
                self.store_viewer.setColumnCount(len(header))
                
                for i in range(0,len(header)): ##Load correct header and copy it to table
                    self.store_viewer. setHorizontalHeaderItem(i,QtWidgets.QTableWidgetItem(header[i]))
                j=0
                for r in self.sohar.difa_cables:
                    item=QtWidgets.QTableWidgetItem(str(r))
                    item.setTextAlignment(128 | 4)
                    self.store_viewer.setItem(j,0, item)
                    j+=1
                    
                j=0
                for r in self.sohar.difa_cables: ## Loop over sohar.cables dict keys
                    for i in range(0,len(self.sohar.difa_cables[r])): ## loop over the list value of the given key
                        item=QtWidgets.QTableWidgetItem(str(self.sohar.difa_cables[r][i]))
                        item.setTextAlignment(128 | 4)
                        self.store_viewer.setItem(j,i+1, item)
                    j+=1

                for i in range(len(header)):
                    self.store_viewer.resizeColumnToContents(i)

                    
            elif self.st_selector.currentText()=='Sohar Used Cables': ## Sohar cables is selected
                header=['Reel no','Galv New','Galv Used','SS New','SS Used','Size']
                self.store_viewer.setRowCount(len(self.sohar.used_cables))
                self.store_viewer.setColumnCount(len(header))
                
                for i in range(0,len(header)): ##Load correct header and copy it to table
                    self.store_viewer. setHorizontalHeaderItem(i,QtWidgets.QTableWidgetItem(header[i]))
                j=0
                for r in self.sohar.used_cables:
                    item=QtWidgets.QTableWidgetItem(str(r))
                    item.setTextAlignment(128 | 4)
                    self.store_viewer.setItem(j,0, item)
                    j+=1
                    
                j=0
                for r in self.sohar.used_cables: ## Loop over sohar.cables dict keys
                    for i in range(0,len(self.sohar.used_cables[r])): ## loop over the list value of the given key
                        item=QtWidgets.QTableWidgetItem(str(self.sohar.used_cables[r][i]))
                        item.setTextAlignment(128 | 4)
                        self.store_viewer.setItem(j,i+1, item)
                    j+=1

                for i in range(len(header)):
                    self.store_viewer.resizeColumnToContents(i)
                        
            elif self.st_selector.currentText()=='Sohar Consumables': ## Sohar consumables is selected
                con_header=['Description','Part Number','1C Code','SN','Quantity']
                self.store_viewer.setRowCount(len(self.sohar.consumables))
                self.store_viewer.setColumnCount(len(con_header))
                for i in range(0,len(con_header)): ##Load correct header and copy it to table
                    self.store_viewer. setHorizontalHeaderItem(i,QtWidgets.QTableWidgetItem(con_header[i]))
                j=0

                for r in self.sohar.consumables: ## Loop over sohar.consumables dict keys
                    for i in range(0,len(self.sohar.consumables[r])):   ## loop over the list value of the given key
                        item=QtWidgets.QTableWidgetItem(str(self.sohar.consumables[r][i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        else:
                            pass
                        self.store_viewer.setItem(j,i, item)
                        
                        
##                        self.store_viewer.setItem(j,i, QtWidgets.QTableWidgetItem(str(self.sohar.consumables[r][i]))) ## Copy list items to table row
                    j+=1
                
                self.store_viewer.resizeColumnToContents(0)


            
            elif self.st_selector.currentText()=='Sohar History': ## Sohar history is selected
                s_header=['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition','Date out','Base','Well']
                self.store_viewer.setRowCount(len(self.sohar.shistory))
                self.store_viewer.setColumnCount(len(self.sohar.shistory[0]))
                
                self.store_viewer. setHorizontalHeaderLabels(s_header)
                j=0
                for L in self.sohar.shistory: ## Loop over sHistory lists
                    for i in range(0,len(L)): ## Loop over the given list
                        item=QtWidgets.QTableWidgetItem(str(L[i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        else:
                            pass
                        self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i,QtWidgets.QTableWidgetItem(str(L[i]))) ## copy list items to table row
                    j+=1 ## increment col number
                
                self.store_viewer.resizeColumnToContents(0)


            elif self.st_selector.currentText()=='Repair History': ## Repair history is selected
                r_header=['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition','Repair Date']
##                self.store_viewer.setRowCount(len(self.sohar.rhistory))
                self.store_viewer.setColumnCount(len(r_header))
                self.store_viewer. setHorizontalHeaderLabels(r_header)
                j=0
                for k in self.sohar.rhistory:
                    for L in self.sohar.rhistory[k]: ## Loop over rHistory lists
                        self.store_viewer.insertRow(j)
                        for i in range(0,len(L)): ## Loop over the given list
                            item=QtWidgets.QTableWidgetItem(str(L[i]))
                            if i!=0:
                                item.setTextAlignment(128 | 4)
                            else:
                                pass
                            self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i,QtWidgets.QTableWidgetItem(str(L[i]))) ## copy list items to table row
                        j+=1 ## increment col number
                for i in range(len(r_header)):
                    self.store_viewer.resizeColumnToContents(i)

                                    
            elif self.st_selector.currentText()=='Booked Sets': ## Booked sets is selected
                b_header=['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition','Base','Well']
                self.store_viewer.setRowCount(len(self.sohar.bookings))
                self.store_viewer.setColumnCount(len(b_header))
                for i in range(0,len(b_header)): ##Load correct header and copy it to table
                    self.store_viewer. setHorizontalHeaderItem(i,QtWidgets.QTableWidgetItem(b_header[i]))
                j=0
                for L in self.sohar.bookings: ## Loop over sohar.bookings lists
                    
                    for i in range(0,len(L)): ## Loop over the given list
                        item=QtWidgets.QTableWidgetItem(str(L[i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        else:
                            pass
                        self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i,QtWidgets.QTableWidgetItem(str(L[i]))) ## copy list items to table row
                    j+=1
                
                self.store_viewer.resizeColumnToContents(0)
            
                        
            elif self.st_selector.currentText()=='Sohar Used equipment': ## Sohar Used Equipment is selected
                b_header=['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition','Base','Well']
                self.store_viewer.setRowCount(len(self.sohar.used))
                self.store_viewer.setColumnCount(len(b_header))
                for i in range(0,len(b_header)): ##Load correct header and copy it to table
                    self.store_viewer. setHorizontalHeaderItem(i,QtWidgets.QTableWidgetItem(b_header[i]))
                j=0
                for r in self.sohar.used: ## Loop over sohar.sp dict keys
                    for i in range(0,len(self.sohar.used[r])): ## Loop over the list values for the given key
                        item=QtWidgets.QTableWidgetItem(str(self.sohar.used[r][i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        else:
                            pass
                        self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i, QtWidgets.QTableWidgetItem(str(self.sohar.used[r][i]))) ## copy list items to table row
                    j+=1 ## increment col number
                
                self.store_viewer.resizeColumnToContents(0)
            elif self.st_selector.currentText()=='Sohar DIFA equipment': ##Min store
                difa_header=['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition','Base','Well']
                self.store_viewer.setRowCount(len(self.sohar.difa))
                self.store_viewer.setColumnCount(len(difa_header))
                self.store_viewer. setHorizontalHeaderLabels(difa_header)
                j=0
                for r in self.sohar.difa: ## Loop over sohar.sp dict keys
                    for i in range(0,len(self.sohar.difa[r])): ## Loop over the list values for the given key
                        item=QtWidgets.QTableWidgetItem(str(self.sohar.difa[r][i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        else:
                            pass
                        self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i, QtWidgets.QTableWidgetItem(str(self.sohar.used[r][i]))) ## copy list items to table row
                    j+=1 ## increment col number
                for i in range(len(difa_header)):
                    self.store_viewer.resizeColumnToContents(i)
                    
            elif self.st_selector.currentText()=='Minimum Store Levels': ##Min store
                min_header=['Description','Part Number','Av Quantity','Min Store required','Intransit order','Intransit QTY','ETA']
                self.store_viewer.setRowCount(len(self.sohar.minStore_map))
                self.store_viewer.setColumnCount(len(min_header))
                self.store_viewer.setHorizontalHeaderLabels(min_header)
                
               
                
                j=0
                for k in self.sohar.minStore_map :
                    
                    if k in self.sohar.orders:
                        a=self.sohar.orders[k][0]
                        b=self.sohar.orders[k][1]
                        c=self.sohar.orders[k][2]
                        
                    else:
                        a=None
                        b=None
                        c=None

                    L=[self.sohar.pn2desc_map()[k],k,self.sohar.av_pn_map()[k],self.sohar.minStore_map[k],a,b,c]
                    print(L)
                    for i in range(0,len(L)):
                        item=QtWidgets.QTableWidgetItem(str(L[i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        if L[2]-L[3]>5:
                            item.setBackground(QtGui.QColor(0,255,0))
                        elif L[2]-L[3]>0 and L[2]-L[3]<=5:
                            item.setBackground(QtGui.QColor(255,255,0))
                        elif L[2]-L[3]<=0:
                            item.setBackground(QtGui.QColor(255,0,0))
                        self.store_viewer.setItem(j,i,item)
                        
                        
                    j+=1
                    
                for i in range(0,len(min_header)):
                    self.store_viewer.resizeColumnToContents(i)
                        
            elif self.st_selector.currentText()=='Cable History':
                cab_header=['Reel Number','Galv New','Galv Used','SS New','SS Used','Size','Update','Arm/Cond','Well','Operation','Date']
                self.store_viewer.setRowCount(len(self.sohar.chistory))
                self.store_viewer.setColumnCount(len(cab_header))
                self.store_viewer.setHorizontalHeaderLabels(cab_header)
                j=0
                for L in self.sohar.chistory: ## Loop over sohar.bookings lists
                    
                    for i in range(0,len(L)): ## Loop over the given list
                        item=QtWidgets.QTableWidgetItem(str(L[i]))
                        
                        item.setTextAlignment(128 | 4)
                        
                        self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i,QtWidgets.QTableWidgetItem(str(L[i]))) ## copy list items to table row
                    j+=1
                for i in range(len(cab_header)):
                    self.store_viewer.resizeColumnToContents(i)
                
            elif self.st_selector.currentText()=='Cable Scrap':
                header=['Description','Part Number','Quantity']
                self.store_viewer.setRowCount(len(self.sohar.scrap))
                self.store_viewer.setColumnCount(len(header))
                self.store_viewer.setHorizontalHeaderLabels(header)

                j=0
                for k in self.sohar.scrap:
                    for i in range(len(header)):
                        item=QtWidgets.QTableWidgetItem(str(self.sohar.scrap[k][i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        self.store_viewer.setItem(j,i, item)
                    j+=1
                for i in range(len(header)):
                    self.store_viewer.resizeColumnToContents(i)
                    
                            
            elif self.st_selector.currentText()=='Active Workorder': ## Active workorder is selected
                w_header=['Description','Part Number','1C Code','Serial Number','Quantity','Date In','Condition',self.WO.get_base(),self.WO.get_name(),self.WO.get_status()]
                self.store_viewer.setRowCount(len(self.WO.DH)+len(self.WO.cables)+len(self.WO.consumables)+1)
                self.store_viewer.setColumnCount(len(w_header))
                for i in range(0,len(w_header)): ##Load correct header and copy it to table
                    self.store_viewer.setHorizontalHeaderItem(i,QtWidgets.QTableWidgetItem(w_header[i]))
                j=0
                for r in self.WO.DH: ##Loop over WO.DH dict keys
                    for i in range(0,len(self.WO.DH[r])): ## Loop over the list values for the given key
                        item=QtWidgets.QTableWidgetItem(str(self.WO.DH[r][i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        
                        self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i, QtWidgets.QTableWidgetItem(str(self.WO.DH[r][i]))) ## copy list items to table row
                    j+=1 ## increment col number
                
                for r in self.WO.cables: ##Loop over WO.cables dict keys
                    for i in range(0,len(self.WO.cables[r])): ## Loop over the list values for the given key
                        item=QtWidgets.QTableWidgetItem(str(self.WO.cables[r][i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        if i ==4:
                            q=self.WO.cables[r][i]
                        
                        
                       
                        self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i, QtWidgets.QTableWidgetItem(str(self.WO.cables[r][i]))) ## copy list items to table row
                    j+=1 ## increment row number
                    
                if len(self.WO.rogue)>0:
                    item=QtWidgets.QTableWidgetItem(str(q-self.WO.rogue[0]))
                    item.setTextAlignment(128 | 4)
                    self.store_viewer.setItem((j-1),4,item )
                    c2=self.sohar.cables[self.WO.rogue[1]].copy()
                    c2[4]=self.WO.rogue[0]
                    c2[3]=self.WO.rogue[2]
                    
                    for i in range(0,len(c2)):
                        item=QtWidgets.QTableWidgetItem(str(c2[i]))
                        if i !=0:
                            item.setTextAlignment(128 | 4)
                            
                        self.store_viewer.setItem(j,i,item ) ## copy list items to table row
                    j+=1
  
                for r in self.WO.consumables: ##Loop over WO.consumables dict keys
                    for i in range(0,len(self.WO.consumables[r])): ## Loop over the list values for the given key
##                        if i==2:
##                            continue ## skip last entry
                        item=QtWidgets.QTableWidgetItem(str(self.WO.consumables[r][i]))
                        if i!=0:
                            item.setTextAlignment(128 | 4)
                        else:
                            pass
                        self.store_viewer.setItem(j,i, item)
##                        self.store_viewer.setItem(j,i, QtWidgets.QTableWidgetItem(str(self.WO.consumables[r][i]))) ## copy list items to table row
                    item=QtWidgets.QTableWidgetItem(str(self.WO.consumables[r][4]))
                    item.setTextAlignment(128 | 4)
                    self.store_viewer.setItem(j,4,item ) ## Make last entry manually to qty col in table
                    j+=1 ## increment col number
                
                self.store_viewer.resizeColumnToContents(0)
                self.store_viewer.resizeColumnToContents(1)
                self.store_viewer.resizeColumnToContents(2)
            self.store_viewer.setSortingEnabled(True)
                
        except:
            pass
        
class Repair_dlg(QtWidgets.QDialog,repair_form.Ui_Dialog):

    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.setModal(0)
        header=['PN','QTY']
        self.sp_table.setHorizontalHeaderLabels(header)
        self.show()
        self.add_pn.clicked.connect(lambda:self.add2Table('PN'))
        self.confirm_rep.clicked.connect(self.retrieve_data)
        self.sub_pn.clicked.connect(lambda: self.remfromTable('PN'))
        
    def retrieve_data(self):
        if self.rep_date.text()!='':
            sn=self.sn_input.text()
            self.sp_list=[]
            
            for i in range(0,self.sp_table.rowCount()):
                self.sp_list.append((str(self.sp_table.item(i,0).text()),int(self.sp_table.item(i,1).text())))
            self.close()
        else:
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Warning)
            msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msgBox.setWindowTitle("Warning")
            msgBox.setText("Please enter repair date.")
            msgBox.exec_()
        

    def remfromTable(self,ID):
        
        a=self.sp_table.currentRow()
        self.sp_table.removeRow(a)
        
        
    def add2Table(self,ID):

        pn=self.pn_input.text()
        qty=self.pn_qty.value()
        count=self.sp_table.rowCount()
        self.sp_table.insertRow(count)
        self.sp_table.setItem(count,0,QtWidgets.QTableWidgetItem(pn))
        self.sp_table.setItem(count,1,QtWidgets.QTableWidgetItem(str(qty)))
        self.sp_table.resizeColumnToContents(0)
        self.sp_table.resizeColumnToContents(1)
        self.pn_input.clear()
            
    

class WO_creator(QtWidgets.QDialog,wo_creator.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.setModal(0)
        self.sn_table.setHorizontalHeaderLabels(['SN','QTY'])
        self.pn_table.setHorizontalHeaderLabels(['PN','QTY'])
        self.show()
        

        self.add_sn.clicked.connect(lambda:self.add2Table('SN'))
        self.add_pn.clicked.connect(lambda:self.add2Table('PN'))
        self.confirm_wo.clicked.connect(self.retrieve_data)
        self.cancel.clicked.connect(self.close)
        self.sub_sn.clicked.connect(lambda: self.remfromTable('SN'))
        self.sub_pn.clicked.connect(lambda: self.remfromTable('PN'))

    def retrieve_data(self):
        self.well_name=self.well.text()
        self.well_name=self.well_name.upper()
        if self.well_name!='':
            self.base_name=self.base.currentText()
            sn=[]
            pn=[]
            for i in range(0,self.sn_table.rowCount()):
                sn.append((self.sn_table.item(i,0).text(),int(self.sn_table.item(i,1).text())))
            for i in range(0,self.pn_table.rowCount()):
                pn.append((self.pn_table.item(i,0).text(),int(self.pn_table.item(i,1).text())))
            self.sn_list=sn
            self.pn_list=pn
            self.close()
        else:
            msgBox = QtWidgets.QMessageBox()
            msgBox.setIcon(QtWidgets.QMessageBox.Warning)
            msgBox.setWindowIcon(QtGui.QIcon('icon2.ico'))
            msgBox.setWindowTitle("Warning")
            msgBox.setText("Please enter well name.")
            msgBox.exec_()
   

    def remfromTable(self,ID):
        if ID =='SN':
            a=self.sn_table.currentRow()
            self.sn_table.removeRow(a)
        else:
            a=self.pn_table.currentRow()
            self.pn_table.removeRow(a)
        
        
    def add2Table(self,ID):
        
        if ID=='SN':
            sn=self.sn_input.text()
            qty=self.sn_qty.value()
            count=self.sn_table.rowCount()
            self.sn_table.insertRow(count)
            self.sn_table.setItem(count,0,QtWidgets.QTableWidgetItem(sn))
            self.sn_table.setItem(count,1,QtWidgets.QTableWidgetItem(str(qty)))
            self.sn_table.resizeColumnToContents(0)
            self.sn_table.resizeColumnToContents(1)
            self.sn_input.clear()
        else:
            pn=self.pn_input.text()
            qty=self.pn_qty.value()
            count=self.pn_table.rowCount()
            self.pn_table.insertRow(count)
            self.pn_table.setItem(count,0,QtWidgets.QTableWidgetItem(pn))
            self.pn_table.setItem(count,1,QtWidgets.QTableWidgetItem(str(qty)))
            self.pn_table.resizeColumnToContents(0)
            self.pn_table.resizeColumnToContents(1)
            self.pn_input.clear()        

class MinStore_Dlg(QtWidgets.QDialog,minStore_dlg.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.setModal(0)
        self.add_pn.clicked.connect(self.pn_add)
        self.sub_pn.clicked.connect(self.pn_sub)
        self.pn_table.setSortingEnabled(True)
        self.show()
        
    def pn_add(self):
        pn=self.pn_input.text()
        qty=self.min_qty.value()
        count=self.pn_table.rowCount()
        if count==0:
            self.pn_table.insertRow(count)
            self.pn_table.setItem(count,0,QtWidgets.QTableWidgetItem(pn))
            self.pn_table.setItem(count,1,QtWidgets.QTableWidgetItem(str(qty)))
            self.pn_table.resizeColumnToContents(0)
            self.pn_table.resizeColumnToContents(1)
            self.pn_input.clear()
        else:
            state=[]
            for i in range(0,count):
                if self.pn_table.item(i,0).text()==pn:
                    state.append(False)
                    
                else:
                    state.append(True)
            if False in state:
                msg=QtWidgets.QMessageBox()
                msg.setIcon(QtWidgets.QMessageBox.Critical)
                msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
                msg.setWindowTitle("Error")
                msg.setText("Part Number already entered. Please check again.")
                msg.exec_()
            else:
                self.pn_table.insertRow(count)
                self.pn_table.setItem(count,0,QtWidgets.QTableWidgetItem(pn))
                self.pn_table.setItem(count,1,QtWidgets.QTableWidgetItem(str(qty)))
                self.pn_table.resizeColumnToContents(0)
                self.pn_table.resizeColumnToContents(1)
                self.pn_input.clear()
     
                    
    def pn_sub(self):
        a=self.pn_table.currentRow()
        self.pn_table.removeRow(a)
        
class MinStore_Dlg2(QtWidgets.QDialog,minStore_dlg2.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.setModal(0)
        self.checkBox.stateChanged.connect(self.checking)
        self.show()
    def checking(self):
        if self.checkBox.isChecked():
            self.lineEdit.clear()
            self.lineEdit.setReadOnly(True)
        else:
            self.lineEdit.setReadOnly(False)

class MinStore_Dlg3(QtWidgets.QDialog,minStore_dlg3.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)

class Validation(QtWidgets.QDialog,validation.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.setModal(0)
        self.show()
class ItemSubDlg(QtWidgets.QDialog,WO_Edit.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.setModal(0)
        self.show()
class ItemAddDlg(QtWidgets.QDialog,WO_Edit.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.setModal(0)
        self.show()
class Returning_cable(QtWidgets.QDialog,Return_cable.Ui_Dialog):
    def __init__(self,sohar,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.sohar=sohar
        self.check=[]
        self.refresh()

        self.return_command.clicked.connect(self.add2return)
        self.remove_command.clicked.connect(self.rem2return)
        self.pushButton.clicked.connect(self.search)
        self.base_select.currentIndexChanged.connect(self.refresh)

    def refresh(self):##Bookmark1
        self.base_view.clear()
        if self.base_select.currentText()=='Lekhwair':
            j=0
            if len(self.sohar.lekh)>0:
                for k in self.sohar.lekh:
                    if type(self.sohar.lekh[k][0])!=str:
                        self.base_view.insertItem(j,k)
                        j+=1
                    else:
                        pass
        else:
            j=0
            if len(self.sohar.nimr)>0:
                for k in self.sohar.nimr:
                    if type(self.sohar.nimr[k][0])!=str:
                        self.base_view.insertItem(j,k)
                        j+=1
                    else:
                        pass
            else:
                pass

    def add2return(self):
        sn=self.base_view.currentItem().text()

        if sn in self.check:
             msg=QtWidgets.QMessageBox()
             msg.setIcon(QtWidgets.QMessageBox.Critical)
             msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
             msg.setWindowTitle("Error")
             msg.setText("Item already in return list.")
             msg.exec_()
             return
        else:
            self.check.append(sn)
            
        sn2=QtWidgets.QTableWidgetItem(sn)
        self.return_view.insertRow(self.return_view.rowCount())
        self.return_view.setItem(self.return_view.rowCount()-1,0,sn2)
        for i in range(self.return_view.rowCount()):
            self.return_view.resizeColumnToContents(i) 
    def rem2return(self):
        a=self.return_view.currentRow()
        if a==-1:
            return
        sn=self.return_view.item(a,0).text()
        for data in self.check:
            if data==sn:
                self.check.remove(data)
        self.return_view.removeRow(a)
        
    def search(self):
        result=[]
        
        if self.sn_filter.text()=='':
            self.refresh()
        else:
            self.base_view.clear()
            if self.base_select.currentText()=='Lekhwair':
                for k in self.sohar.lekh:
                    if type(self.sohar.lekh[k][0])!=str:
                        if self.sn_filter.text() in k:
                            result.append(k)
                j=0  
                for entry in result:
                    self.base_view.insertItem(j,entry)
                    j+=1
            elif self.base_select.currentText()=='Nimr':
                for k in self.sohar.nimr:
                    if type(self.sohar.nimr[k][0])!=str:
                        if self.sn_filter.text() in k:
                            result.append(k)
                j=0  
                for entry in result:
                    self.base_view.insertItem(j,entry)
                    j+=1
class Returning_set(QtWidgets.QDialog,return_wo.Ui_Dialog):
    def __init__(self,sohar,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.sohar=sohar
        self.check=[]
        self.refresh()
        

        self.return_command.clicked.connect(self.add2return)
        self.remove_command.clicked.connect(self.rem2return)
        self.pushButton.clicked.connect(self.search)
        self.base_select.currentIndexChanged.connect(self.refresh)

    def refresh(self):
        
        
        if self.base_select.currentText()=='Lekhwair':

            self.base_view.setRowCount(len(self.sohar.lekh))
            
            j=0
            if len(self.sohar.lekh)>0:
                for k in self.sohar.lekh:
                    if type(self.sohar.lekh[k][0])==str:
                        entry=[self.sohar.lekh[k][0],self.sohar.lekh[k][3]]
                        for i in range(len(entry)):
                            item=QtWidgets.QTableWidgetItem(entry[i])
                            if i==1:
                                item.setTextAlignment(128 | 4)
                            self.base_view.setItem(j,i,item)
                        j+=1
                    else:
                        pass
                for i in range(self.base_view.rowCount()):
                    self.base_view.resizeColumnToContents(i)
            else:
                pass
        else:
            if len(self.sohar.nimr)>0:
                self.base_view.setRowCount(len(self.sohar.nimr))
                j=0
                for k in self.sohar.nimr:
                    if type(self.sohar.nimr[k][0])==str:
                        entry=[self.sohar.nimr[k][0],self.sohar.nimr[k][3]]
                        for i in range(len(entry)):
                            item=QtWidgets.QTableWidgetItem(entry[i])
                            if i>0:
                                item.setTextAlignment(128 | 4)
                            
                            self.base_view.setItem(j,i,item)
                            
                        j+=1
                    else:
                        pass
                for i in range(self.base_view.rowCount()):
                    self.base_view.resizeColumnToContents(i)
            else:
                pass
        
    def add2return(self):
        if self.base_view.currentRow()==-1:
            return
        sn=self.base_view.item(self.base_view.currentRow(),1).text()
        if self.base_select.currentText()=='Lekhwair':
            data=[self.sohar.lekh[sn][0],self.sohar.lekh[sn][3]]
        else:
            data=[self.sohar.nimr[sn][0],self.sohar.nimr[sn][3]]
        if data in self.check:
             msg=QtWidgets.QMessageBox()
             msg.setIcon(QtWidgets.QMessageBox.Critical)
             msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
             msg.setWindowTitle("Error")
             msg.setText("Item already in return list.")
             msg.exec_()
             return
        else:
            self.check.append(data)
            
                
        self.return_view.insertRow(self.return_view.rowCount())
        for i in range(len(data)):
            item=QtWidgets.QTableWidgetItem(data[i])
            if i>0:
                item.setTextAlignment(128 | 4)
            self.return_view.setItem(self.return_view.rowCount()-1,i,item)
        for i in range(self.return_view.rowCount()):
            self.return_view.resizeColumnToContents(i) 
                
    def rem2return(self):
        a=self.return_view.currentRow()
        if a==-1:
            return
        sn=self.return_view.item(a,1).text()
        for data in self.check:
            if data[1]==sn:
                self.check.remove(data)
        self.return_view.removeRow(a)
        
    def search(self):
        result=[]
        
        if self.sn_filter.text()=='':
            self.refresh()
        else:
            self.base_view.clear()
            if self.base_select.currentText()=='Lekhwair':
                self.base_view.setHorizontalHeaderLabels(['Description','Serial No'])
                for k in self.sohar.lekh:
                    if self.sn_filter.text() in self.sohar.lekh[k][3]:
                        result.append([self.sohar.lekh[k][0],self.sohar.lekh[k][3]])
                j=0  
                for entry in result:
                    for i in range(len(entry)):
                        item=QtWidgets.QTableWidgetItem(entry[i])
                        if i>0:
                            item.setTextAlignment(128 | 4)
                        
                        self.base_view.setItem(j,i,item)
                    j+=1
            elif self.base_select.currentText()=='Nimr':
                self.base_view.setHorizontalHeaderLabels(['Description','Serial No'])
                for k in self.sohar.nimr:
                    if self.sn_filter.text() in self.sohar.nimr[k][3]:
                        result.append([self.sohar.nimr[k][0],self.sohar.nimr[k][3]])
                j=0  
                for entry in result:
                    for i in range(len(entry)):
                        item=QtWidgets.QTableWidgetItem(entry[i])
                        if i>0:
                            item.setTextAlignment(128 | 4)
                        
                        self.base_view.setItem(j,i,item)
                    j+=1
        for i in range(self.base_view.rowCount()):
            self.base_view.resizeColumnToContents(i)

class clearing_difa_cables(QtWidgets.QDialog,clear_difa2.Ui_Dialog)  :
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        
        
class Clearing_difa(QtWidgets.QDialog,clear_difa.Ui_Dialog):
    def __init__(self,sohar,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.sohar=sohar
        self.check=[]
        self.refresh()

        self.clear_command.clicked.connect(self.add2return)
        self.keep_command.clicked.connect(self.rem2return)
        self.search.clicked.connect(self.searching)
        

    def refresh(self):
        
        self.base_view.setRowCount(len(self.sohar.difa))
        
        j=0
        if len(self.sohar.difa)>0:
            for k in self.sohar.difa:
                entry=[self.sohar.difa[k][0],self.sohar.difa[k][3]]
                for i in range(len(entry)):
                    item=QtWidgets.QTableWidgetItem(entry[i])
                    if i==1:
                        item.setTextAlignment(128 | 4)
                    self.base_view.setItem(j,i,item)
                j+=1
            for i in range(self.base_view.rowCount()):
                self.base_view.resizeColumnToContents(i)
        else:
            pass
        
        
    def add2return(self):
        if self.base_view.currentRow()==-1:
            return
        sn=self.base_view.item(self.base_view.currentRow(),1).text()
        data=[self.sohar.difa[sn][0],self.sohar.difa[sn][3]]
        
        if data in self.check:
             msg=QtWidgets.QMessageBox()
             msg.setIcon(QtWidgets.QMessageBox.Critical)
             msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
             msg.setWindowTitle("Error")
             msg.setText("Item already in return list.")
             msg.exec_()
             return
        else:
            self.check.append(data)
            
                
        self.cleared_view.insertRow(self.cleared_view.rowCount())
        for i in range(len(data)):
            item=QtWidgets.QTableWidgetItem(data[i])
            if i>0:
                item.setTextAlignment(128 | 4)
            self.cleared_view.setItem(self.cleared_view.rowCount()-1,i,item)
        for i in range(self.cleared_view.rowCount()):
            self.cleared_view.resizeColumnToContents(i) 
                
    def rem2return(self):
        a=self.cleared_view.currentRow()
        if a==-1:
            return
        sn=self.cleared_view.item(a,1).text()
        for data in self.check:
            if data[1]==sn:
                self.check.remove(data)
        self.cleared_view.removeRow(a)
        
    def searching(self):
        result=[]
        
        if self.sn_filter.text()=='':
            self.refresh()
        else:
            self.base_view.clear()
            self.base_view.setHorizontalHeaderLabels(['Description','Serial No'])
            for k in self.sohar.difa:
                if self.sn_filter.text() in self.sohar.difa[k][3]:
                    result.append([self.sohar.difa[k][0],self.sohar.difa[k][3]])
            j=0  
            for entry in result:
                for i in range(len(entry)):
                    item=QtWidgets.QTableWidgetItem(entry[i])
                    if i>0:
                        item.setTextAlignment(128 | 4)
                    
                    self.base_view.setItem(j,i,item)
                j+=1
            
        for i in range(self.base_view.rowCount()):
            self.base_view.resizeColumnToContents(i)
                
                        
                
            

class Sending_set(QtWidgets.QDialog,send_wo.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
class Batch_dlg(QtWidgets.QDialog,new_batch.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.loadFile.clicked.connect(self.load_batch)

    def load_batch(self):
        
        filename=QtWidgets.QFileDialog.getOpenFileName(QtWidgets.QDialog(),'Open File', 
                                                       os.getenv('C:\\Users\\HP\\AppData\\Local\\Programs\\Python\\Python36-32\\Python exercises'))
        self.batchFile.setText(filename[0])
        self.file=filename[0]

class Reassignment(QtWidgets.QDialog,WO_reassign.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
class scrapping(QtWidgets.QDialog,cable_scrap.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.setModal(0)
        self.show()

class reelUpdateDlg(QtWidgets.QDialog,reel_update.Ui_Dialog):
    def __init__(self,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        self.setModal(0)
        self.show()

class wells_list(QtWidgets.QDialog,wells.Ui_Dialog):
    def __init__(self,wells,parent=None):
        QtWidgets.QDialog.__init__(self,parent)
        self.setupUi(self)
        wells.sort()
        self.wells=wells
        for i,w in enumerate(self.wells):
            self.well_list.insertItem(i,w)

        self.search.clicked.connect(self.filter_wells)

    def filter_wells(self):
        if self.filter.text()=='':
            for i,w in enumerate(self.wells):
                self.well_list.insertItem(i,w)
            
        result=[]
        for w in self.wells:
            if self.filter.text().lower() in w.lower():
                result.append(w)
        self.well_list.clear()
        for i,w in enumerate(result):
            self.well_list.insertItem(i,w)
        
        
if lic.isExpired():
    app=QtWidgets.QApplication([])
    msg=QtWidgets.QMessageBox()
    msg.setIcon(QtWidgets.QMessageBox.Critical)
    msg.setWindowIcon(QtGui.QIcon('icon2.ico'))
    msg.setWindowTitle("Error")
    msg.setText("Internal maintenance needed.")
    msg.exec_()
else:
    app=QtWidgets.QApplication([])
##    pixmap = QtGui.QPixmap("image.png")
    pixmap = QtGui.QPixmap("newlogo.png")
    splash = QtWidgets.QSplashScreen(pixmap)
    splash.show()
    app.processEvents()
    time.sleep(3)
    s=Dashboard()
    s.showMaximized()
    splash.finish(s)
    app.exec_()
